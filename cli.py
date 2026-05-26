#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import contextlib
import datetime as dt
import fcntl
import hashlib
import json
import os
import re
import shlex
import shutil
import signal
import sqlite3
import subprocess
import sys
import time
import urllib.error
import urllib.request
import uuid
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SERVICE_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_PATH = SERVICE_DIR / "config" / "defaults.json"


def _load_config() -> dict[str, Any]:
    config_path = Path(os.environ.get("INVOICE_CHECK_CONFIG", DEFAULT_CONFIG_PATH)).expanduser()
    data: dict[str, Any] = {}
    if config_path.is_file():
        data = json.loads(config_path.read_text(encoding="utf-8"))
    env_map = {
        "adb_path": "INVOICE_CHECK_ADB_PATH",
        "vlm_api": "INVOICE_CHECK_VLM_API",
        "vlm_model": "INVOICE_CHECK_VLM_MODEL",
        "vlm_api_key": "INVOICE_CHECK_VLM_API_KEY",
        "paddle_vl_api": "INVOICE_CHECK_PADDLE_VL_API",
        "paddle_vl_model": "INVOICE_CHECK_PADDLE_VL_MODEL",
        "paddle_vl_api_key": "INVOICE_CHECK_PADDLE_VL_API_KEY",
        "remote_dir": "INVOICE_CHECK_REMOTE_DIR",
        "remote_image": "INVOICE_CHECK_REMOTE_IMAGE",
        "cleanup_gallery_for_invoice": "INVOICE_CHECK_CLEANUP_GALLERY_FOR_INVOICE",
        "ocr_cli_path": "INVOICE_CHECK_OCR_CLI_PATH",
        "qr_precheck_mode": "INVOICE_CHECK_QR_PRECHECK_MODE",
        "qr_wechat_model_dir": "INVOICE_CHECK_QR_WECHAT_MODEL_DIR",
        "runtime_dir": "INVOICE_CHECK_RUNTIME_DIR",
        "queue_db_path": "INVOICE_CHECK_QUEUE_DB_PATH",
        "adb_serial": "INVOICE_CHECK_ADB_SERIAL",
        "phone_resource_id": "INVOICE_CHECK_PHONE_RESOURCE_ID",
        "phone_display_name": "INVOICE_CHECK_PHONE_DISPLAY_NAME",
    }
    for key, env_name in env_map.items():
        if os.environ.get(env_name):
            data[key] = os.environ[env_name]
    numeric_envs: dict[str, tuple[str, type]] = {
        "vlm_temperature": ("INVOICE_CHECK_VLM_TEMPERATURE", float),
        "vlm_top_p": ("INVOICE_CHECK_VLM_TOP_P", float),
        "vlm_top_k": ("INVOICE_CHECK_VLM_TOP_K", int),
        "vlm_max_tokens": ("INVOICE_CHECK_VLM_MAX_TOKENS", int),
        "vlm_timeout": ("INVOICE_CHECK_VLM_TIMEOUT", float),
        "paddle_vl_temperature": ("INVOICE_CHECK_PADDLE_VL_TEMPERATURE", float),
        "paddle_vl_max_tokens": ("INVOICE_CHECK_PADDLE_VL_MAX_TOKENS", int),
        "paddle_vl_timeout": ("INVOICE_CHECK_PADDLE_VL_TIMEOUT", float),
        "phone_lock_wait_seconds": ("INVOICE_CHECK_PHONE_LOCK_WAIT_SECONDS", float),
        "queue_wait_timeout_seconds": ("INVOICE_CHECK_QUEUE_WAIT_TIMEOUT_SECONDS", float),
        "max_steps": ("INVOICE_CHECK_MAX_STEPS", int),
        "batch_attempts": ("INVOICE_CHECK_BATCH_ATTEMPTS", int),
        "max_same_state_visits": ("INVOICE_CHECK_MAX_SAME_STATE_VISITS", int),
        "max_same_action_repeats": ("INVOICE_CHECK_MAX_SAME_ACTION_REPEATS", int),
        "max_h5_loading_waits": ("INVOICE_CHECK_MAX_H5_LOADING_WAITS", int),
        "qr_preprocess_canvas_size": ("INVOICE_CHECK_QR_PREPROCESS_CANVAS_SIZE", int),
        "qr_preprocess_qr_max_size": ("INVOICE_CHECK_QR_PREPROCESS_QR_MAX_SIZE", int),
    }
    for key, (env_name, caster) in numeric_envs.items():
        if os.environ.get(env_name):
            data[key] = caster(os.environ[env_name])
    bool_envs = {
        "vlm_enable_thinking": "INVOICE_CHECK_VLM_ENABLE_THINKING",
        "failed_replay_enabled": "INVOICE_CHECK_FAILED_REPLAY_ENABLED",
        "qr_precheck_enabled": "INVOICE_CHECK_QR_PRECHECK_ENABLED",
        "qr_precheck_block_unreadable": "INVOICE_CHECK_QR_PRECHECK_BLOCK_UNREADABLE",
        "qr_preprocess_enabled": "INVOICE_CHECK_QR_PREPROCESS_ENABLED",
        "qr_preprocess_require_decode": "INVOICE_CHECK_QR_PREPROCESS_REQUIRE_DECODE",
        "qr_crop_fallback_enabled": "INVOICE_CHECK_QR_CROP_FALLBACK_ENABLED",
    }
    for key, env_name in bool_envs.items():
        if os.environ.get(env_name):
            raw = os.environ[env_name].strip().lower()
            data[key] = raw in {"1", "true", "yes", "on"}
    return data


CONFIG = _load_config()
ADB = Path(CONFIG.get("adb_path", "adb")).expanduser()
VLM_API = CONFIG.get("vlm_api", "http://127.0.0.1:8081/v1/chat/completions")
VLM_MODEL = CONFIG.get("vlm_model", "qwen-vl")
VLM_API_KEY = str(CONFIG.get("vlm_api_key") or "").strip()
PADDLE_VL_API = CONFIG.get("paddle_vl_api", "http://127.0.0.1:8090/v1/chat/completions")
PADDLE_VL_MODEL = CONFIG.get("paddle_vl_model", "paddleocr-vl")
PADDLE_VL_API_KEY = str(CONFIG.get("paddle_vl_api_key") or "").strip()
REMOTE_DIR = CONFIG.get("remote_dir", "/sdcard/Pictures/invoice-check")
REMOTE_IMAGE = CONFIG.get("remote_image", "qr_current.png")
CLEANUP_GALLERY_FOR_INVOICE = bool(CONFIG.get("cleanup_gallery_for_invoice", True))
OCR_CLI_PATH = Path(CONFIG.get("ocr_cli_path") or (Path(__file__).resolve().parent.parent / "ocr-cli" / "bin" / "ocr-cli"))
ADB_SERIAL = str(CONFIG.get("adb_serial") or "").strip() or None
VLM_TEMPERATURE = float(CONFIG.get("vlm_temperature", 0.01))
VLM_TOP_P = float(CONFIG.get("vlm_top_p", 0.8))
VLM_TOP_K = int(CONFIG.get("vlm_top_k", 20))
VLM_MAX_TOKENS = int(CONFIG.get("vlm_max_tokens", 512))
VLM_TIMEOUT = float(CONFIG.get("vlm_timeout", 30))
VLM_ENABLE_THINKING = bool(CONFIG.get("vlm_enable_thinking", False))
PADDLE_VL_TEMPERATURE = float(CONFIG.get("paddle_vl_temperature", 0))
PADDLE_VL_MAX_TOKENS = int(CONFIG.get("paddle_vl_max_tokens", 768))
PADDLE_VL_TIMEOUT = float(CONFIG.get("paddle_vl_timeout", 15))
EMIT_PROGRESS_EVENTS = bool(CONFIG.get("emit_progress_events", False))
RUNTIME_DIR = Path(CONFIG.get("runtime_dir", "/tmp/agentd-invoice-check-cli")).expanduser()
QUEUE_DB_PATH = Path(CONFIG.get("queue_db_path") or (RUNTIME_DIR / "queue.sqlite3")).expanduser()
PHONE_RESOURCE_ID = str(CONFIG.get("phone_resource_id", "android_phone:invoice-default"))
PHONE_DISPLAY_NAME = str(CONFIG.get("phone_display_name", "发票查验手机"))
PHONE_LOCK_WAIT_SECONDS = float(CONFIG.get("phone_lock_wait_seconds", 7200))
QUEUE_WAIT_TIMEOUT_SECONDS = float(CONFIG.get("queue_wait_timeout_seconds", 7200))
MAX_SAME_STATE_VISITS = int(CONFIG.get("max_same_state_visits", 3))
MAX_SAME_ACTION_REPEATS = int(CONFIG.get("max_same_action_repeats", 2))
MAX_H5_LOADING_WAITS = int(CONFIG.get("max_h5_loading_waits", 2))
QR_PRECHECK_ENABLED = bool(CONFIG.get("qr_precheck_enabled", True))
QR_PRECHECK_BLOCK_UNREADABLE = bool(CONFIG.get("qr_precheck_block_unreadable", True))
QR_PRECHECK_MODE = str(CONFIG.get("qr_precheck_mode") or "fast_positive_only").strip().lower()
if os.environ.get("INVOICE_CHECK_QR_PRECHECK_BLOCK_UNREADABLE") and not os.environ.get("INVOICE_CHECK_QR_PRECHECK_MODE"):
    QR_PRECHECK_MODE = "block_unreadable" if QR_PRECHECK_BLOCK_UNREADABLE else "fast_positive_only"
QR_PREPROCESS_ENABLED = bool(CONFIG.get("qr_preprocess_enabled", True))
QR_PREPROCESS_REQUIRE_DECODE = bool(CONFIG.get("qr_preprocess_require_decode", True))
QR_PREPROCESS_CANVAS_SIZE = int(CONFIG.get("qr_preprocess_canvas_size", 1000))
QR_PREPROCESS_QR_MAX_SIZE = int(CONFIG.get("qr_preprocess_qr_max_size", 860))
_wechat_model_dir_raw = str(CONFIG.get("qr_wechat_model_dir") or (SERVICE_DIR / "models" / "wechat_qrcode"))
QR_WECHAT_MODEL_DIR = Path(_wechat_model_dir_raw)
if not QR_WECHAT_MODEL_DIR.is_absolute():
    QR_WECHAT_MODEL_DIR = SERVICE_DIR / QR_WECHAT_MODEL_DIR
QR_CROP_FALLBACK_ENABLED = bool(CONFIG.get("qr_crop_fallback_enabled", True))
STOP_REQUESTED = False


class CancelledError(RuntimeError):
    pass


class ModelDecisionError(RuntimeError):
    pass


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def _handle_stop_signal(signum: int, _frame: Any) -> None:
    global STOP_REQUESTED
    STOP_REQUESTED = True
    emit_progress("cancel signal received", signal=signum)


signal.signal(signal.SIGTERM, _handle_stop_signal)
signal.signal(signal.SIGINT, _handle_stop_signal)


@dataclass
class Candidate:
    id: int
    label_hint: str
    box: list[int]

    @property
    def center(self) -> tuple[int, int]:
        x1, y1, x2, y2 = self.box
        return (x1 + x2) // 2, (y1 + y2) // 2

    def as_json(self) -> dict[str, Any]:
        x, y = self.center
        return {"id": self.id, "label_hint": self.label_hint, "box": self.box, "center": [x, y]}


def resource_slug(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "resource"


def ensure_runtime_dirs() -> None:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    (RUNTIME_DIR / "locks").mkdir(parents=True, exist_ok=True)
    (RUNTIME_DIR / "cancel").mkdir(parents=True, exist_ok=True)


class PhoneLock:
    def __init__(self, *, task_id: str | None = None, wait_seconds: float | None = None) -> None:
        ensure_runtime_dirs()
        self.task_id = task_id
        self.wait_seconds = PHONE_LOCK_WAIT_SECONDS if wait_seconds is None else wait_seconds
        self.slug = resource_slug(PHONE_RESOURCE_ID)
        self.lock_path = RUNTIME_DIR / "locks" / f"{self.slug}.lock"
        self.lease_path = RUNTIME_DIR / "locks" / f"{self.slug}.lease.json"
        self._fh: Any = None

    def __enter__(self) -> "PhoneLock":
        deadline = time.monotonic() + max(0, self.wait_seconds)
        self.lock_path.parent.mkdir(parents=True, exist_ok=True)
        self._fh = self.lock_path.open("a+")
        while True:
            try:
                fcntl.flock(self._fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                self._write_lease("active")
                emit_progress("phone resource lock acquired", resource_id=PHONE_RESOURCE_ID, task_id=self.task_id)
                return self
            except BlockingIOError:
                if self.wait_seconds <= 0 or time.monotonic() >= deadline:
                    owner = read_json_file(self.lease_path)
                    raise RuntimeError(
                        json.dumps(
                            {
                                "error": "device_busy",
                                "resource_id": PHONE_RESOURCE_ID,
                                "owner": owner,
                            },
                            ensure_ascii=False,
                        )
                    )
                emit_progress("phone resource busy; waiting", resource_id=PHONE_RESOURCE_ID, task_id=self.task_id)
                sleep_with_cancel(2)

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        try:
            self._write_lease("released")
        finally:
            if self._fh is not None:
                fcntl.flock(self._fh.fileno(), fcntl.LOCK_UN)
                self._fh.close()
                self._fh = None
            with contextlib.suppress(FileNotFoundError):
                self.lease_path.unlink()
            emit_progress("phone resource lock released", resource_id=PHONE_RESOURCE_ID, task_id=self.task_id)

    def touch(self) -> None:
        self._write_lease("active")

    def _write_lease(self, state: str) -> None:
        payload = {
            "resource_id": PHONE_RESOURCE_ID,
            "display_name": PHONE_DISPLAY_NAME,
            "state": state,
            "pid": os.getpid(),
            "task_id": self.task_id,
            "adb_serial": ADB_SERIAL,
            "heartbeat_at": utc_now(),
        }
        self.lease_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json_file(path: Path) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def cancel_file_for_task(task_id: str) -> Path:
    ensure_runtime_dirs()
    return RUNTIME_DIR / "cancel" / f"{task_id}.cancel"


def is_cancelled(cancel_file: str | None = None) -> bool:
    if STOP_REQUESTED:
        return True
    if cancel_file and Path(cancel_file).is_file():
        return True
    env_cancel = os.environ.get("AGENTD_CANCEL_FILE") or os.environ.get("INVOICE_CHECK_CANCEL_FILE")
    return bool(env_cancel and Path(env_cancel).is_file())


def check_cancelled(cancel_file: str | None = None) -> None:
    if is_cancelled(cancel_file):
        raise CancelledError("invoice check task cancelled")


def sleep_with_cancel(seconds: float, cancel_file: str | None = None) -> None:
    end = time.monotonic() + max(0, seconds)
    while time.monotonic() < end:
        check_cancelled(cancel_file)
        time.sleep(min(0.5, end - time.monotonic()))
    check_cancelled(cancel_file)


def adb_base_cmd() -> list[str]:
    cmd = [str(ADB)]
    if ADB_SERIAL:
        cmd.extend(["-s", ADB_SERIAL])
    return cmd


def run(
    cmd: list[str],
    *,
    capture: bool = False,
    check: bool = True,
    timeout: float = 30,
    cancel_file: str | None = None,
    stdout_file: Any | None = None,
) -> subprocess.CompletedProcess[str]:
    check_cancelled(cancel_file)
    stdout: Any
    stderr: Any
    if stdout_file is not None:
        stdout = stdout_file
        stderr = subprocess.PIPE if capture else subprocess.DEVNULL
    elif capture:
        stdout = subprocess.PIPE
        stderr = subprocess.PIPE
    else:
        stdout = None
        stderr = None
    text_mode = stdout_file is None
    proc = subprocess.Popen(cmd, text=text_mode, stdout=stdout, stderr=stderr)
    started = time.monotonic()
    while proc.poll() is None:
        if timeout is not None and time.monotonic() - started > timeout:
            terminate_process(proc)
            raise subprocess.TimeoutExpired(cmd, timeout)
        try:
            check_cancelled(cancel_file)
        except CancelledError:
            terminate_process(proc)
            raise
        time.sleep(0.2)
    out, err = proc.communicate()
    result = subprocess.CompletedProcess(cmd, proc.returncode, out or "", err or "")
    if check and result.returncode != 0:
        raise subprocess.CalledProcessError(result.returncode, cmd, output=result.stdout, stderr=result.stderr)
    return result


def terminate_process(proc: subprocess.Popen[str]) -> None:
    if proc.poll() is not None:
        return
    with contextlib.suppress(ProcessLookupError):
        proc.terminate()
    try:
        proc.wait(timeout=2)
    except subprocess.TimeoutExpired:
        with contextlib.suppress(ProcessLookupError):
            proc.kill()
        with contextlib.suppress(subprocess.TimeoutExpired):
            proc.wait(timeout=2)


def adb(args: list[str], *, capture: bool = False, check: bool = True, timeout: float = 30) -> subprocess.CompletedProcess[str]:
    check_cancelled()
    cmd = [*adb_base_cmd(), *args]
    try:
        return run(cmd, capture=capture, check=check, timeout=timeout)
    except subprocess.TimeoutExpired:
        recover_adb_server()
        return run(cmd, capture=capture, check=check, timeout=timeout)


def recover_adb_server() -> None:
    # ADB occasionally wedges when prior commands are interrupted. Best effort only.
    subprocess.run([str(ADB), "kill-server"], text=True, capture_output=True, timeout=5, check=False)
    time.sleep(0.5)
    subprocess.run([str(ADB), "start-server"], text=True, capture_output=True, timeout=10, check=False)
    time.sleep(0.5)


def adb_shell(command: str, *, capture: bool = False, check: bool = True, timeout: float = 30) -> subprocess.CompletedProcess[str]:
    # Use one shell string so Android shell handles quoting and globs consistently.
    return adb(["shell", command], capture=capture, check=check, timeout=timeout)


def get_size() -> tuple[int, int]:
    proc = adb(["shell", "wm", "size"], capture=True)
    line = proc.stdout.strip().splitlines()[-1]
    size = line.split(":")[-1].strip()
    width, height = size.split("x")
    return int(width), int(height)


def screenshot(path: Path) -> None:
    check_cancelled()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("wb") as f:
        run([*adb_base_cmd(), "exec-out", "screencap", "-p"], check=True, stdout_file=f, timeout=20)


def tap(x: int, y: int) -> None:
    adb(["shell", "input", "tap", str(x), str(y)], timeout=10)


def keyevent(code: int) -> None:
    adb(["shell", "input", "keyevent", str(code)], timeout=10)


def swipe(x1: int, y1: int, x2: int, y2: int, duration_ms: int = 300) -> None:
    adb([
        "shell",
        "input",
        "swipe",
        str(x1),
        str(y1),
        str(x2),
        str(y2),
        str(duration_ms),
    ], timeout=10)


def delete_remote_media_store_records() -> None:
    where = f"_data LIKE '{REMOTE_DIR}/%'"
    adb_shell(
        f"content delete --uri content://media/external/images/media --where {shlex.quote(where)}",
        timeout=10,
        check=False,
    )


def cleanup_gallery_for_invoice() -> None:
    """Keep the dedicated verification phone gallery from leaking old images.

    WeChat's picker opens "All images", not only REMOTE_DIR. For this dedicated
    verification phone, clear common image locations before staging each invoice
    and after each attempt so the picker has one effective candidate.
    """
    if not CLEANUP_GALLERY_FOR_INVOICE:
        return
    for directory, maxdepth in [
        ("/sdcard/Pictures", 3),
        ("/sdcard/DCIM/Camera", 1),
        ("/sdcard/Download", 1),
    ]:
        adb_shell(
            " ".join(
                [
                    f"mkdir -p {shlex.quote(directory)};",
                    f"find {shlex.quote(directory)} -mindepth 1 -maxdepth {maxdepth} -type f",
                    r"\( -iname '*.jpg' -o -iname '*.jpeg' -o -iname '*.png' -o -iname '*.webp' \)",
                    "-delete",
                ]
            ),
            timeout=20,
            check=False,
        )
    media_where = (
        "_data LIKE '/sdcard/Pictures/%' OR _data LIKE '/storage/emulated/0/Pictures/%' "
        "OR _data LIKE '/sdcard/DCIM/Camera/%' OR _data LIKE '/storage/emulated/0/DCIM/Camera/%' "
        "OR _data LIKE '/sdcard/Download/%' OR _data LIKE '/storage/emulated/0/Download/%'"
    )
    adb_shell(
        f"content delete --uri content://media/external/images/media --where {shlex.quote(media_where)}",
        timeout=20,
        check=False,
    )


def remote_invoice_dir_file_count() -> int | None:
    proc = adb_shell(
        f"mkdir -p {REMOTE_DIR!r}; find {REMOTE_DIR!r} -maxdepth 1 -type f | wc -l",
        capture=True,
        timeout=10,
        check=False,
    )
    try:
        return int(proc.stdout.strip().splitlines()[-1])
    except (IndexError, ValueError):
        return None


def validate_remote_album_isolated(expected_remote: str) -> dict[str, Any]:
    count = remote_invoice_dir_file_count()
    payload = {
        "remote_dir": REMOTE_DIR,
        "expected_remote": expected_remote,
        "file_count": count,
        "isolated": count == 1,
    }
    if count != 1:
        raise RuntimeError(f"phone media isolation failed: expected 1 file in {REMOTE_DIR}, got {count}")
    return payload


def stage_image(local_image: Path, remote_name: str | None = None) -> str:
    if not local_image.is_file():
        raise FileNotFoundError(local_image)
    cleanup_gallery_for_invoice()
    adb_shell(f"mkdir -p {REMOTE_DIR!r}", timeout=10)
    adb_shell(f"find {REMOTE_DIR!r} -maxdepth 1 -type f -delete", timeout=10)
    delete_remote_media_store_records()
    remote = f"{REMOTE_DIR}/{remote_name or REMOTE_IMAGE}"
    adb(["push", str(local_image), remote], timeout=20)
    adb([
        "shell",
        "am",
        "broadcast",
        "-a",
        "android.intent.action.MEDIA_SCANNER_SCAN_FILE",
        "-d",
        f"file://{remote}",
    ], timeout=10)
    validate_remote_album_isolated(remote)
    return remote


def cleanup_phone_album() -> None:
    cleanup_gallery_for_invoice()
    adb_shell(f"mkdir -p {REMOTE_DIR!r}", timeout=10, check=False)
    adb_shell(f"find {REMOTE_DIR!r} -maxdepth 1 -type f -delete", timeout=10, check=False)
    delete_remote_media_store_records()
    adb([
        "shell",
        "am",
        "broadcast",
        "-a",
        "android.intent.action.MEDIA_SCANNER_SCAN_FILE",
        "-d",
        f"file://{REMOTE_DIR}/",
    ], timeout=10, check=False)


def is_textish_pixel(rgb: tuple[int, int, int]) -> bool:
    r, g, b = rgb
    dark = r < 120 and g < 120 and b < 120
    green = g > 110 and r < 90 and b < 140
    gray = 120 <= r <= 215 and abs(r - g) < 16 and abs(g - b) < 16
    white = r > 225 and g > 225 and b > 225
    red = r > 180 and g < 90 and b < 100
    return dark or green or gray or white or red


def connected_components_text_boxes(image: Image.Image) -> list[list[int]]:
    width, height = image.size
    pix = image.load()
    visited: set[tuple[int, int]] = set()
    comps: list[list[int]] = []
    y_min = 70
    y_max = min(height, 900)
    for y in range(y_min, y_max):
        for x in range(width):
            if (x, y) in visited or not is_textish_pixel(pix[x, y]):
                continue
            stack = [(x, y)]
            visited.add((x, y))
            xs: list[int] = []
            ys: list[int] = []
            for px, py in stack:
                xs.append(px)
                ys.append(py)
                for nx in (px - 1, px, px + 1):
                    for ny in (py - 1, py, py + 1):
                        if (
                            0 <= nx < width
                            and y_min <= ny < y_max
                            and (nx, ny) not in visited
                            and is_textish_pixel(pix[nx, ny])
                        ):
                            visited.add((nx, ny))
                            stack.append((nx, ny))
            if len(xs) >= 8:
                x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
                if (x2 - x1) >= 2 and (y2 - y1) >= 4:
                    comps.append([x1, y1, x2, y2])

    lines: list[list[Any]] = []
    for x1, y1, x2, y2 in sorted(comps, key=lambda b: (b[1], b[0])):
        if y2 - y1 > 90 or x2 - x1 > width * 0.9:
            continue
        placed = False
        cy = (y1 + y2) / 2
        for line in lines:
            lx1, ly1, lx2, ly2, items = line
            lcy = (ly1 + ly2) / 2
            if abs(cy - lcy) < 18:
                items.append([x1, y1, x2, y2])
                line[0] = min(lx1, x1)
                line[1] = min(ly1, y1)
                line[2] = max(lx2, x2)
                line[3] = max(ly2, y2)
                placed = True
                break
        if not placed:
            lines.append([x1, y1, x2, y2, [[x1, y1, x2, y2]]])

    boxes: list[list[int]] = []
    for _lx1, _ly1, _lx2, _ly2, items in lines:
        run: list[list[int]] = []
        last: list[int] | None = None
        for item in sorted(items, key=lambda b: b[0]):
            if last is not None and item[0] - last[2] > 42:
                if run:
                    boxes.append(_padded_run_box(run, width, height))
                run = []
            run.append(item)
            last = item
        if run:
            boxes.append(_padded_run_box(run, width, height))
    return boxes


def _padded_run_box(run: list[list[int]], width: int, height: int) -> list[int]:
    x1 = min(c[0] for c in run)
    y1 = min(c[1] for c in run)
    x2 = max(c[2] for c in run)
    y2 = max(c[3] for c in run)
    return [max(0, x1 - 16), max(0, y1 - 14), min(width, x2 + 16), min(height, y2 + 14)]


def color_button_boxes(image: Image.Image) -> list[tuple[str, list[int]]]:
    width, height = image.size
    pix = image.load()
    boxes: list[tuple[str, list[int]]] = []
    rows: list[tuple[int, int, int]] = []
    for y in range(height):
        xs: list[int] = []
        for x in range(width):
            r, g, b = pix[x, y]
            if g > 145 and r < 90 and b < 140:
                xs.append(x)
        if len(xs) > 45:
            rows.append((min(xs), y, max(xs)))
    if rows:
        groups: list[list[tuple[int, int, int]]] = []
        for row in rows:
            if groups and row[1] - groups[-1][-1][1] <= 2:
                groups[-1].append(row)
            else:
                groups.append([row])
        for group in groups:
            if len(group) < 8:
                continue
            x1 = min(r[0] for r in group)
            y1 = min(r[1] for r in group)
            x2 = max(r[2] for r in group)
            y2 = max(r[1] for r in group) + 1
            boxes.append(("green_button_or_selected_tab", [x1, y1, x2, y2]))
    return boxes


def is_scan_line_box(box: list[int], width: int, height: int) -> bool:
    x1, y1, x2, y2 = box
    box_width = x2 - x1
    box_height = y2 - y1
    center_y = (y1 + y2) / 2
    return (
        box_height <= 22
        and box_width >= max(80, width * 0.16)
        and height * 0.28 <= center_y <= height * 0.72
    )


def dark_pixel_ratio(image: Image.Image) -> float:
    width, height = image.size
    x_start = int(width * 0.08)
    x_end = int(width * 0.92)
    y_start = int(height * 0.16)
    y_end = int(height * 0.84)
    total = 0
    dark = 0
    pix = image.load()
    for y in range(y_start, y_end, 8):
        for x in range(x_start, x_end, 8):
            r, g, b = pix[x, y]
            total += 1
            if r + g + b < 135:
                dark += 1
    return dark / total if total else 0.0


def looks_like_wechat_camera_scan_page(
    image: Image.Image,
    page_hint: str = "",
    green_boxes: list[tuple[str, list[int]]] | None = None,
) -> bool:
    compact = re.sub(r"\s+", "", page_hint or "").lower()
    scan_markers = ["我的二维码", "轻触照亮", "扫一扫", "扫码页", "扫描二维码", "scanqrcode"]
    if any(marker in compact for marker in scan_markers):
        return True
    width, height = image.size
    boxes = green_boxes if green_boxes is not None else color_button_boxes(image)
    has_scan_line = any(is_scan_line_box(box, width, height) for _label, box in boxes)
    return has_scan_line and dark_pixel_ratio(image) >= 0.42


def camera_album_icon_box(width: int, height: int) -> list[int]:
    return [
        max(0, int(width * 0.78)),
        max(0, int(height * 0.88)),
        min(width, int(width * 0.99)),
        min(height, int(height * 0.99)),
    ]


def anchor_boxes(
    width: int,
    height: int,
    page_hint: str = "",
    *,
    is_camera_scan_page: bool = False,
) -> list[tuple[str, list[int]]]:
    compact = re.sub(r"\s+", "", page_hint or "").lower()
    has_fiscal = "财政票据" in compact
    has_search_context = any(marker in compact for marker in ["搜索", "最近", "search", "recent"])
    has_invoice_check = any(marker in compact for marker in ["票据查验", "电子票号查验", "电子票据代码"])
    has_service_context = has_fiscal and any(
        marker in compact for marker in ["服务号", "公众号", "服务", "菜单", "提供", "票据查验"]
    )
    has_camera_scan_context = any(marker in compact for marker in ["扫一扫", "扫码页", "扫描二维码", "我的二维码", "轻触照亮"])
    has_album_context = any(marker in compact for marker in ["相册", "图库", "选择图片", "照片"])
    has_dialog_context = any(
        marker in compact for marker in ["提示", "确认", "确定", "无效", "查验异常", "超过当日", "次数"]
    )

    if is_camera_scan_page:
        return [
            ("top_left_close_or_back", [0, 75, 85, 155]),
            ("camera_album_icon", camera_album_icon_box(width, height)),
        ]

    anchors = [
        ("top_left_close_or_back", [0, 75, 85, 155]),
        ("top_right_menu", [620, 75, 715, 155]),
        ("wechat_home_search_icon", [540, 75, 635, 155]),
        ("wechat_search_input", [70, 80, 690, 175]),
    ]
    if has_fiscal and has_search_context:
        anchors.extend(
            [
                ("wechat_recent_fiscal_ticket_search", [20, 330, 280, 420]),
                ("wechat_search_result_service_card", [20, 270, 705, 595]),
            ]
        )
    if has_fiscal and has_invoice_check:
        anchors.append(("wechat_search_result_ticket_check_button", [170, 485, 430, 575]))
    if has_invoice_check:
        anchors.extend(
            [
                ("scan_tab_expected_area", [460, 165, 620, 235]),
                ("service_tab_expected_area", [170, 570, 290, 665]),
            ]
        )
    if has_service_context:
        anchors.extend(
            [
                ("service_account_bottom_menu_ticket_check", [500, 1475, min(width, 715), min(height, 1630)]),
                ("service_popup_ticket_check_row", [0, 1480, width, min(height, 1630)]),
            ]
        )
    if has_camera_scan_context:
        anchors.append(("camera_album_icon", camera_album_icon_box(width, height)))
    if has_album_context:
        anchors.append(("album_first_image_slot", [178, 150, 365, 340]))
    if has_dialog_context:
        anchors.append(("dialog_confirm_button", [150, 860, 570, 970]))
    return anchors


def build_candidates(
    image_path: Path,
    annotated_path: Path,
    extra_boxes: list[tuple[str, list[int]]] | None = None,
    page_hint: str = "",
) -> list[Candidate]:
    image = Image.open(image_path).convert("RGB")
    width, height = image.size
    is_launcher = looks_like_android_launcher(page_hint)
    raw: list[tuple[str, list[int]]] = []
    if extra_boxes:
        raw.extend(extra_boxes)
    if not is_launcher:
        green_boxes = color_button_boxes(image)
        is_camera_scan_page = looks_like_wechat_camera_scan_page(image, page_hint, green_boxes)
        for box in connected_components_text_boxes(image):
            x1, y1, x2, y2 = box
            if x2 - x1 >= 18 and y2 - y1 >= 12:
                raw.append(("cv_text_or_ui_region", box))
        for label, box in green_boxes:
            if is_camera_scan_page and is_scan_line_box(box, width, height):
                continue
            raw.append((label, box))
        raw.extend(anchor_boxes(width, height, page_hint=page_hint, is_camera_scan_page=is_camera_scan_page))

    deduped: list[tuple[str, list[int]]] = []
    for label, box in raw:
        x1, y1, x2, y2 = box
        if x2 <= x1 or y2 <= y1:
            continue
        area = (x2 - x1) * (y2 - y1)
        if area > width * height * 0.35:
            continue
        if any(_iou(box, existing) > 0.82 for _label, existing in deduped):
            continue
        deduped.append((label, box))

    priority: list[tuple[str, list[int]]] = []
    regular: list[tuple[str, list[int]]] = []
    for item in deduped:
        label, box = item
        if (
            label.startswith("paddle_launcher_app:微信")
            or label.startswith("paddle_ocr:微信")
            or label.startswith("heuristic_launcher_app:微信")
        ):
            priority.append(item)
        else:
            regular.append(item)

    # Keep high-confidence OCR app targets first; fill the rest top-to-bottom.
    regular = sorted(regular, key=lambda item: (item[1][1], item[1][0]))
    deduped = priority + regular
    deduped = deduped[:60]
    candidates = [Candidate(i + 1, label, box) for i, (label, box) in enumerate(deduped)]

    out = image.copy()
    draw = ImageDraw.Draw(out)
    colors = ["red", "blue", "orange", "purple", "lime", "cyan", "magenta", "yellow"]
    for c in candidates:
        color = colors[(c.id - 1) % len(colors)]
        x1, y1, x2, y2 = c.box
        draw.rectangle(c.box, outline=color, width=4)
        draw.rectangle([x1, max(0, y1 - 28), x1 + 58, y1], fill=color)
        draw.text((x1 + 5, max(0, y1 - 25)), str(c.id), fill="white")
    out.save(annotated_path)
    image.close()
    return candidates


def looks_like_android_launcher(page_hint: str) -> bool:
    if not page_hint:
        return False
    launcher_markers = [
        "日历",
        "钱包",
        "游戏中心",
        "小米视频",
        "米家",
        "文件管理",
        "小米社区",
        "拼多多",
        "淘宝",
    ]
    workflow_markers = [
        "财政票据",
        "票据查验",
        "扫码查验",
        "服务通知",
        "相册",
        "图库",
        "腾讯新闻",
        "订阅号",
        "搜索",
        "聊天",
        "通讯录",
        "发现",
    ]
    wechat_internal_markers = [
        "我的二维码",
        "添加朋友",
        "收照片",
        "收红包",
        "财政票据",
        "票据查验",
        "扫码查验",
        "服务通知",
        "相册",
        "图库",
        "腾讯新闻",
        "订阅号",
        "通讯录",
        "发现",
    ]
    if any(marker in page_hint for marker in workflow_markers):
        return False
    if "微信" in page_hint and not any(marker in page_hint for marker in wechat_internal_markers):
        return True
    return sum(1 for marker in launcher_markers if marker in page_hint) >= 3


def simplified_wechat_launcher_boxes(page_hint: str, width: int, height: int) -> list[tuple[str, list[int]]]:
    if "微信" not in page_hint:
        return []
    if any(marker in page_hint for marker in ["我的二维码", "添加朋友", "收照片", "财政票据", "票据查验", "扫码查验", "腾讯新闻", "订阅号", "通讯录", "发现"]):
        return []
    # The test phone launcher is intentionally simplified to a single WeChat
    # icon near the upper-left. PaddleOCR may return only plain text without
    # LOC tokens, so provide a deterministic tappable app cell fallback.
    return [("heuristic_launcher_app:微信", [40, 80, min(width, 180), min(height, 260)])]


def _iou(a: list[int], b: list[int]) -> float:
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    if ix2 <= ix1 or iy2 <= iy1:
        return 0.0
    inter = (ix2 - ix1) * (iy2 - iy1)
    area_a = (ax2 - ax1) * (ay2 - ay1)
    area_b = (bx2 - bx1) * (by2 - by1)
    return inter / float(area_a + area_b - inter)


def maybe_ask_paddle_vl(image_path: Path) -> str:
    """Optional OCR-ish page summary via PaddleOCR-VL-compatible OpenAI endpoint."""
    try:
        b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
        payload = {
            "model": PADDLE_VL_MODEL,
            "temperature": PADDLE_VL_TEMPERATURE,
            "max_tokens": PADDLE_VL_MAX_TOKENS,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "Read the business-relevant UI text from this phone screenshot. "
                                "Ignore the status bar, battery/network icons, repeated speaker icons, and decorative symbols. "
                                "Return concise plain text lines only; do not emit coordinates, LOC tokens, markdown, or JSON."
                            ),
                        },
                        {"type": "image_url", "image_url": {"url": "data:image/png;base64," + b64}},
                    ],
                }
            ],
        }
        resp = post_json(PADDLE_VL_API, payload, timeout=PADDLE_VL_TIMEOUT, api_key=PADDLE_VL_API_KEY)
        raw = resp.get("choices", [{}])[0].get("message", {}).get("content", "")
        return sanitize_paddle_hint(raw)
    except Exception as exc:  # noqa: BLE001 - best-effort helper for MVP traces
        return f"paddle_vl_unavailable: {type(exc).__name__}: {exc}"


def sanitize_paddle_hint(raw: str, *, max_chars: int = 1200) -> str:
    text = re.sub(r"<\|LOC_\d+\|>", "", str(raw or ""))
    lines: list[str] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"([^\w\s])(?:\s*\1){4,}", r"\1", line)
        line = re.sub(r"\b([A-Za-z0-9])(?:\s+\1){8,}\b", r"\1", line)
        line = re.sub(r"([0-9])(?:\s+\1){8,}", r"\1", line)
        if len(line) > 120:
            line = line[:120]
        if line and line not in lines:
            lines.append(line)
        if sum(len(item) + 1 for item in lines) >= max_chars:
            break
    return "\n".join(lines)[:max_chars]


def looks_like_post_scan_return_to_service_page(text: str) -> bool:
    compact = _compact_text(text, limit=1200).replace(" ", "")
    if "财政票据" not in compact or "票据查验" not in compact:
        return False
    return any(marker in compact for marker in ["轻触屏幕继续扫描", "继续扫描", "重新扫描"])


def paddle_text_boxes(raw: str, width: int, height: int) -> list[tuple[str, list[int]]]:
    """Parse PaddleOCR-VL location-token text into candidate boxes.

    The vLLM PaddleOCR-VL endpoint returns lines like:
    ``微信<|LOC_118|><|LOC_720|>...`` where LOC coordinates are normalized
    to a 0..1000 grid. Convert them to screenshot pixels.
    """
    boxes: list[tuple[str, list[int]]] = []
    pattern = re.compile(r"(.+?)((?:<\|LOC_\d+\|>){8})")
    loc_pattern = re.compile(r"<\|LOC_(\d+)\|>")
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        match = pattern.search(line)
        if not match:
            continue
        text = match.group(1).strip()
        locs = [int(x) for x in loc_pattern.findall(match.group(2))]
        if len(locs) != 8:
            continue
        xs = locs[0::2]
        ys = locs[1::2]
        x1 = round(min(xs) * width / 1000)
        x2 = round(max(xs) * width / 1000)
        y1 = round(min(ys) * height / 1000)
        y2 = round(max(ys) * height / 1000)
        if x2 <= x1 or y2 <= y1:
            continue
        padded = [max(0, x1 - 18), max(0, y1 - 16), min(width, x2 + 18), min(height, y2 + 16)]
        boxes.append((f"paddle_ocr:{text}", padded))
        if text in {"微信", "WeChat"}:
            # Launcher app labels sit below the icon. Add a larger tappable cell
            # centered above the label, so tapping opens the app instead of just
            # hitting the text baseline.
            cx = (x1 + x2) // 2
            cell = [max(0, cx - 75), max(0, y1 - 150), min(width, cx + 75), min(height, y2 + 20)]
            boxes.append(("paddle_launcher_app:微信", cell))
    return boxes


def call_vlm(
    image_path: Path,
    candidates: list[Candidate],
    paddle_hint: str,
    step: int,
    stop_at: str,
    task_phase: str = "verify",
) -> dict[str, Any]:
    b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
    cand_json = [c.as_json() for c in candidates]
    instruction = {
        "task": (
            "Leave the current invoice verification flow and reach any safe stable screen."
            if task_phase == "return_to_entry"
            else "Verify the currently staged invoice QR code in WeChat fiscal invoice verification."
        ),
        "task_phase": task_phase,
        "current_step": step,
        "allowed_actions": [
            "tap_candidate",
            "back",
            "wait",
            "swipe_left",
            "swipe_right",
            "home",
            "stop_success",
            "stop_user_action_required",
            "stop_failed",
        ],
        "policy": [
            "Use candidate_id instead of raw coordinates.",
            "Always choose the smallest/specific candidate that directly covers the target, not a broad container.",
            f"Current stop_at mode is {stop_at}.",
            f"Current task_phase is {task_phase}.",
            "If task_phase is return_to_entry, your only goal is to leave the current invoice verification flow and reach a safe stable screen.",
            "If task_phase is return_to_entry, safe stable screens include WeChat home/chat list, the 财政票据 chat page, Android home, or any non-invoice page.",
            "If task_phase is return_to_entry, do not search for the service account, do not tap 票据查验, and do not start a new verification.",
            "If task_phase is return_to_entry and you already see a safe stable screen, choose stop_success immediately.",
            "If task_phase is return_to_entry and the screen is invoice detail/result, 票据查验 form, camera scan page, album picker, or a business popup, choose back or a confirm button to leave it.",
            "If stop_at is album and the photo picker/album page is visible, your action MUST be stop_success with candidate_id null. Do not tap any photo or QR candidate.",
            "A Xiaomi account expired credential banner/toast on the Android launcher is unrelated to this task. Ignore it or wait briefly; do not stop_user_action_required for that banner alone.",
            "Canonical entry strategy: the normal task entrance is the 财政票据 service account main/chat page in WeChat. From there open the bottom 票据查验 service menu.",
            "If the screen is not WeChat, not Android launcher, and not a manual security/permission/login blocker, choose home to return to the Android launcher, then open WeChat and navigate back to 财政票据.",
            "If inside WeChat but not already in 财政票据 or an invoice-check child page, navigate by WeChat search to 财政票据 before continuing.",
            "Hard rule for the current test phone: Android Home returns to a simple launcher page with only WeChat. On Android launcher/home screen, if WeChat is not clearly visible, choose home once or wait for banners to disappear before swiping.",
            "On Android launcher/home screen, only tap WeChat when a candidate label_hint explicitly contains paddle_launcher_app:微信 or paddle_ocr:微信. Otherwise prefer home/wait first; use swipe only after Home has failed to reveal WeChat.",
            "Page relationship map:",
            "0. Android launcher/home screen -> find and tap the WeChat/微信 app icon only when the candidate clearly contains the WeChat green two-bubble icon or label 微信. If WeChat is not clearly visible, use swipe_left or swipe_right to change launcher page. Never guess by grid position.",
            "1. WeChat home/chat list -> tap search icon.",
            "2. WeChat search page -> if recent/search suggestion 财政票据 is visible, tap it.",
            "3. WeChat search results for 财政票据 -> tap the 财政票据 service account or its 票据查验 button.",
            "4. 财政票据 service account chat/home -> tap the bottom 票据查验 menu item.",
            "5. Fiscal invoice H5 loading/header page -> wait until the form or scan tab appears.",
            "6. 票据查验 form page -> tap 扫码查验.",
            "7. WeChat camera scan page -> tap album/gallery.",
            "8. Photo picker/album page -> select unique QR image, unless stop_at is album.",
            "9. After QR selection -> wait for result; success page shows 票据详情/票据信息/金额合计/查验次数.",
            "Back behavior: search page back returns to WeChat home; search results back returns to search page; service account/chat back returns to search results or WeChat; H5 detail back returns to check form; camera back returns to check form; album back returns to camera.",
            "Important: after tapping 票据查验 on the service account, the pjcy.mof.gov.cn H5 page may show a bare loading/header page first. In task_phase verify, if pjcy.mof.gov.cn is visible but form controls or 扫码查验 tab are not visible yet, choose wait. Do not tap top-left close/back during loading.",
            "Launcher behavior: this phone is configured so Home should reveal the intended WeChat-only launcher page. Do not tap random app icons unless the icon is likely WeChat/微信.",
            "If on 财政票据 service account home, open 服务/服务号提供的服务 then choose 票据查验.",
            "If on 票据查验 form, choose 扫码查验.",
            "If on WeChat camera scan page, choose album/gallery.",
            "If in photo picker, choose the only QR/image thumbnail.",
            "If a fiscal invoice landing/header page is still loading and there is no clear scan/form control, wait instead of tapping random menu or album anchors.",
            "If task_phase is verify and invoice detail/票据信息/查验次数/金额合计 is visible, stop_success.",
            "If 查验异常/server transient appears, report result_status=server_transient. The host supervisor owns retry/stop policy; do not keep looping by yourself.",
            "If phone is locked, WeChat not logged in, permission prompt, or captcha/manual security appears, stop_user_action_required.",
        ],
        "paddle_vl_hint": paddle_hint,
        "candidates": cand_json,
        "response_schema": {
            "action": "tap_candidate|back|wait|swipe_left|swipe_right|home|stop_success|stop_user_action_required|stop_failed",
            "candidate_id": "number or null",
            "screen_state": "short string",
            "result_status": "in_progress|success|server_transient|user_action_required|failed",
            "reason": "short Chinese explanation",
            "confidence": "0..1",
        },
    }
    payload = {
        "model": VLM_MODEL,
        "temperature": VLM_TEMPERATURE,
        "top_p": VLM_TOP_P,
        "top_k": VLM_TOP_K,
        "min_p": 0.0,
        "repeat_penalty": 1.0,
        "max_tokens": VLM_MAX_TOKENS,
        "chat_template_kwargs": {"enable_thinking": VLM_ENABLE_THINKING},
        "messages": [
            {
                "role": "system",
                "content": "You are the main controller for an Android phone invoice-check task. /no_think Return exactly one JSON object. No markdown. Do not output raw coordinates.",
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(instruction, ensure_ascii=False) + "\n/no_think",
                    },
                    {"type": "image_url", "image_url": {"url": "data:image/png;base64," + b64}},
                ],
            },
        ],
    }
    resp = post_json(VLM_API, payload, timeout=VLM_TIMEOUT, api_key=VLM_API_KEY)
    message = resp.get("choices", [{}])[0].get("message", {})
    finish_reason = resp.get("choices", [{}])[0].get("finish_reason")
    content = str(message.get("content") or "")
    if not content.strip():
        reasoning = str(message.get("reasoning_content") or "")[:500]
        raise ModelDecisionError(
            f"empty VLM content; finish_reason={finish_reason}; reasoning_content={reasoning!r}"
        )
    try:
        return parse_json_object(content)
    except json.JSONDecodeError as exc:
        raise ModelDecisionError(f"invalid VLM JSON content: {content[:500]!r}") from exc


def should_home_recover(decision: dict[str, Any]) -> bool:
    text = " ".join(
        str(decision.get(key, "")) for key in ("screen_state", "reason", "result_status", "action")
    ).lower()
    manual_blockers = [
        "locked",
        "锁屏",
        "permission",
        "权限",
        "captcha",
        "验证码",
        "security",
        "安全",
        "login",
        "登录",
        "credential",
        "凭证",
    ]
    if any(marker in text for marker in manual_blockers):
        return False
    recover_markers = [
        "not wechat",
        "non-wechat",
        "other app",
        "wrong app",
        "非微信",
        "其他应用",
        "错误应用",
        "文件管理",
        "file manager",
        "variflight",
        "飞常准",
    ]
    return any(marker in text for marker in recover_markers)


def is_ignorable_system_banner(decision: dict[str, Any]) -> bool:
    text = " ".join(str(decision.get(key, "")) for key in ("screen_state", "reason", "result_status"))
    return "小米账号" in text and ("登录凭证" in text or "凭证失效" in text or "重新登录" in text)


def post_json(url: str, payload: dict[str, Any], *, timeout: float, api_key: str = "") -> dict[str, Any]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = urllib.request.Request(url, data=data, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def parse_json_object(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped)
        stripped = re.sub(r"\s*```$", "", stripped)
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", stripped, re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def queue_connect() -> sqlite3.Connection:
    ensure_runtime_dirs()
    QUEUE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(QUEUE_DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    init_queue_db(conn)
    return conn


def init_queue_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS invoice_check_tasks (
            invoice_task_id TEXT PRIMARY KEY,
            session_id TEXT,
            user_id TEXT,
            title TEXT,
            input_dir TEXT NOT NULL,
            output_dir TEXT NOT NULL,
            status TEXT NOT NULL,
            priority INTEGER NOT NULL DEFAULT 100,
            queued_at TEXT NOT NULL,
            started_at TEXT,
            finished_at TEXT,
            cancelled_at TEXT,
            summary_path TEXT,
            error_message TEXT,
            metadata_json TEXT NOT NULL DEFAULT '{}'
        );

        CREATE INDEX IF NOT EXISTS idx_invoice_check_tasks_queue
        ON invoice_check_tasks(status, priority, queued_at);
        """
    )
    conn.commit()


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    if row is None:
        return None
    data = dict(row)
    if "metadata_json" in data:
        try:
            data["metadata"] = json.loads(data.pop("metadata_json") or "{}")
        except json.JSONDecodeError:
            data["metadata"] = {}
    return data


def submit_queue_task(args: argparse.Namespace) -> dict[str, Any]:
    task_id = args.invoice_task_id or str(uuid.uuid4())
    input_dir = str(Path(args.input_dir).resolve())
    output_dir = str(Path(args.output_dir).resolve())
    now = utc_now()
    metadata = {
        "source": "invoice-check-cli submit",
        "phone_resource_id": PHONE_RESOURCE_ID,
    }
    with queue_connect() as conn:
        conn.execute(
            """
            INSERT INTO invoice_check_tasks (
                invoice_task_id, session_id, user_id, title, input_dir, output_dir,
                status, priority, queued_at, metadata_json
            )
            VALUES (?, ?, ?, ?, ?, ?, 'queued', ?, ?, ?)
            """,
            (
                task_id,
                args.session_id,
                args.user_id,
                args.title or "发票查验任务",
                input_dir,
                output_dir,
                int(args.priority),
                now,
                json.dumps(metadata, ensure_ascii=False),
            ),
        )
        conn.commit()
    return {
        "success": True,
        "status": "queued",
        "invoice_task_id": task_id,
        "resource_id": PHONE_RESOURCE_ID,
        "input_dir": input_dir,
        "output_dir": output_dir,
        "queue_db_path": str(QUEUE_DB_PATH),
    }


def fetch_next_queue_task(conn: sqlite3.Connection) -> dict[str, Any] | None:
    conn.execute("BEGIN IMMEDIATE")
    row = conn.execute(
        """
        SELECT * FROM invoice_check_tasks
        WHERE status = 'queued'
        ORDER BY priority ASC, queued_at ASC
        LIMIT 1
        """
    ).fetchone()
    if row is None:
        conn.commit()
        return None
    now = utc_now()
    conn.execute(
        """
        UPDATE invoice_check_tasks
        SET status = 'running', started_at = ?, error_message = NULL
        WHERE invoice_task_id = ?
        """,
        (now, row["invoice_task_id"]),
    )
    conn.commit()
    updated = conn.execute(
        "SELECT * FROM invoice_check_tasks WHERE invoice_task_id = ?",
        (row["invoice_task_id"],),
    ).fetchone()
    return row_to_dict(updated)


def update_queue_task(task_id: str, **fields: Any) -> None:
    allowed = {
        "status",
        "started_at",
        "finished_at",
        "cancelled_at",
        "summary_path",
        "error_message",
        "metadata_json",
    }
    keys = [key for key in fields if key in allowed]
    if not keys:
        return
    sql = ", ".join(f"{key} = ?" for key in keys)
    values = [fields[key] for key in keys]
    values.append(task_id)
    with queue_connect() as conn:
        conn.execute(f"UPDATE invoice_check_tasks SET {sql} WHERE invoice_task_id = ?", values)
        conn.commit()


def get_queue_task(task_id: str) -> dict[str, Any] | None:
    with queue_connect() as conn:
        row = conn.execute(
            "SELECT * FROM invoice_check_tasks WHERE invoice_task_id = ?",
            (task_id,),
        ).fetchone()
        return row_to_dict(row)


def queue_status_payload() -> dict[str, Any]:
    lease = read_json_file(RUNTIME_DIR / "locks" / f"{resource_slug(PHONE_RESOURCE_ID)}.lease.json")
    with queue_connect() as conn:
        active = conn.execute(
            """
            SELECT * FROM invoice_check_tasks
            WHERE status IN ('running', 'cancel_requested')
            ORDER BY started_at ASC
            LIMIT 1
            """
        ).fetchone()
        queued_rows = conn.execute(
            """
            SELECT invoice_task_id, session_id, user_id, title, queued_at, priority
            FROM invoice_check_tasks
            WHERE status = 'queued'
            ORDER BY priority ASC, queued_at ASC
            """
        ).fetchall()
        recent_rows = conn.execute(
            """
            SELECT invoice_task_id, title, status, queued_at, started_at, finished_at, summary_path, error_message
            FROM invoice_check_tasks
            ORDER BY queued_at DESC
            LIMIT 20
            """
        ).fetchall()
    active_task = row_to_dict(active)
    if active_task or lease:
        status = "busy"
    elif queued_rows:
        status = "queued"
    else:
        status = "idle"
    return {
        "success": True,
        "resource_id": PHONE_RESOURCE_ID,
        "display_name": PHONE_DISPLAY_NAME,
        "status": status,
        "adb_serial": ADB_SERIAL,
        "queue_db_path": str(QUEUE_DB_PATH),
        "active": active_task,
        "lease": lease,
        "queued": len(queued_rows),
        "queue": [dict(row) for row in queued_rows],
        "recent": [dict(row) for row in recent_rows],
    }


def request_cancel(task_id: str) -> dict[str, Any]:
    task = get_queue_task(task_id)
    if task is None:
        return {"success": False, "error": "task_not_found", "invoice_task_id": task_id}
    now = utc_now()
    cancel_file_for_task(task_id).write_text(now + "\n", encoding="utf-8")
    if task["status"] == "queued":
        update_queue_task(task_id, status="cancelled", cancelled_at=now, finished_at=now)
        status = "cancelled"
    elif task["status"] == "running":
        update_queue_task(task_id, status="cancel_requested", cancelled_at=now)
        status = "cancel_requested"
    else:
        status = task["status"]
    return {"success": True, "invoice_task_id": task_id, "status": status}


def discover_qr_tasks(input_dir: Path) -> list[dict[str, Any]]:
    allowed = {".png", ".jpg", ".jpeg", ".webp"}
    tasks: list[dict[str, Any]] = []
    for path in sorted(input_dir.iterdir()):
        if path.is_file() and path.suffix.lower() in allowed:
            tasks.append({"id": safe_task_id(path), "qr_path": str(path.resolve()), "source_name": path.name})
    return tasks


def safe_task_id(path: Path) -> str:
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", path.stem).strip("._")
    return stem or f"qr_{abs(hash(path.name))}"


def precheck_qr_image(path: Path) -> dict[str, Any]:
    if not QR_PRECHECK_ENABLED or QR_PRECHECK_MODE == "disabled":
        return {"enabled": False, "mode": QR_PRECHECK_MODE, "status": "skipped"}
    if not path.is_file():
        return {"enabled": True, "mode": QR_PRECHECK_MODE, "status": "missing", "message": "QR image file is missing."}
    try:
        import cv2  # type: ignore
    except Exception as exc:  # noqa: BLE001 - optional dependency
        return {
            "enabled": True,
            "mode": QR_PRECHECK_MODE,
            "status": "unavailable",
            "message": f"OpenCV QRCodeDetector unavailable: {type(exc).__name__}: {exc}",
        }
    image = cv2.imread(str(path))
    if image is None:
        return {"enabled": True, "mode": QR_PRECHECK_MODE, "status": "unreadable", "message": "OpenCV could not read the image file."}
    detector = cv2.QRCodeDetector()
    data, points, _straight = detector.detectAndDecode(image)
    if data:
        return {
            "enabled": True,
            "mode": QR_PRECHECK_MODE,
            "status": "decoded",
            "payload_hash": hashlib.sha256(data.encode("utf-8")).hexdigest()[:16],
            "has_points": points is not None,
        }
    return {
        "enabled": True,
        "mode": QR_PRECHECK_MODE,
        "status": "unreadable",
        "message": "No QR payload decoded by OpenCV; continuing whole-image scan path.",
        "blocking": QR_PRECHECK_MODE == "block_unreadable",
    }


def _point_bbox(points: list[tuple[float, float]], image_size: tuple[int, int], *, padding: int = 60) -> list[int]:
    width, height = image_size
    xs = [p[0] for p in points if p[0] == p[0]]
    ys = [p[1] for p in points if p[1] == p[1]]
    if not xs or not ys:
        return [0, 0, 0, 0]
    return [
        max(0, int(min(xs)) - padding),
        max(0, int(min(ys)) - padding),
        min(width, int(max(xs)) + padding),
        min(height, int(max(ys)) + padding),
    ]


def _normalize_qr_crop(crop: Image.Image, output_path: Path) -> None:
    """Create a phone-scan-friendly QR image with large quiet zone."""
    canvas_size = max(320, int(QR_PREPROCESS_CANVAS_SIZE))
    qr_max = max(160, min(canvas_size - 80, int(QR_PREPROCESS_QR_MAX_SIZE)))
    source = crop.convert("RGB")
    scale = min(qr_max / max(1, source.width), qr_max / max(1, source.height), 1.0 if max(source.size) >= qr_max else 99.0)
    new_size = (max(1, int(source.width * scale)), max(1, int(source.height * scale)))
    resized = source.resize(new_size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", (canvas_size, canvas_size), "white")
    canvas.paste(resized, ((canvas_size - resized.width) // 2, (canvas_size - resized.height) // 2))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path)


def _qr_candidate_record(
    *,
    engine: str,
    source_path: Path,
    image: Image.Image,
    bbox: list[int],
    payload: str,
    output_dir: Path,
    index: int,
) -> dict[str, Any]:
    x1, y1, x2, y2 = bbox
    if x2 <= x1 or y2 <= y1:
        raise ValueError(f"invalid QR bbox: {bbox}")
    crop = image.crop((x1, y1, x2, y2))
    raw_path = output_dir / f"{index:02d}_{engine}_raw.png"
    normalized_path = output_dir / f"{index:02d}_{engine}_best_qr.png"
    output_dir.mkdir(parents=True, exist_ok=True)
    crop.save(raw_path)
    _normalize_qr_crop(crop, normalized_path)
    return {
        "engine": engine,
        "source": str(source_path),
        "raw_image": str(raw_path),
        "image": str(normalized_path),
        "bbox": bbox,
        "size": [x2 - x1, y2 - y1],
        "decoded": bool(payload),
        "payload_hash": hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16] if payload else "",
    }


def _detect_qr_with_zxing(path: Path, output_dir: Path, *, start_index: int = 1) -> list[dict[str, Any]]:
    try:
        import zxingcpp  # type: ignore
    except Exception:
        return []
    image = Image.open(path).convert("RGB")
    try:
        results = zxingcpp.read_barcodes(image, formats=zxingcpp.BarcodeFormat.QRCode)
    except Exception:
        image.close()
        return []
    candidates: list[dict[str, Any]] = []
    for offset, result in enumerate(results):
        payload = str(getattr(result, "text", "") or "")
        if QR_PREPROCESS_REQUIRE_DECODE and not payload:
            continue
        position = getattr(result, "position", None)
        if not position:
            continue
        points = [
            (float(position.top_left.x), float(position.top_left.y)),
            (float(position.top_right.x), float(position.top_right.y)),
            (float(position.bottom_right.x), float(position.bottom_right.y)),
            (float(position.bottom_left.x), float(position.bottom_left.y)),
        ]
        bbox = _point_bbox(points, image.size, padding=60)
        try:
            candidates.append(
                _qr_candidate_record(
                    engine="zxing",
                    source_path=path,
                    image=image,
                    bbox=bbox,
                    payload=payload,
                    output_dir=output_dir,
                    index=start_index + offset,
                )
            )
        except ValueError:
            continue
    image.close()
    return candidates


def _wechat_model_paths() -> tuple[Path, Path, Path, Path] | None:
    detect_proto = QR_WECHAT_MODEL_DIR / "detect.prototxt"
    detect_model = QR_WECHAT_MODEL_DIR / "detect.caffemodel"
    sr_proto = QR_WECHAT_MODEL_DIR / "sr.prototxt"
    sr_model = QR_WECHAT_MODEL_DIR / "sr.caffemodel"
    if all(path.is_file() for path in (detect_proto, detect_model, sr_proto, sr_model)):
        return detect_proto, detect_model, sr_proto, sr_model
    return None


def _detect_qr_with_wechat(path: Path, output_dir: Path, *, start_index: int = 1) -> list[dict[str, Any]]:
    model_paths = _wechat_model_paths()
    if not model_paths:
        return []
    try:
        import cv2  # type: ignore
    except Exception:
        return []
    ctor = getattr(cv2, "wechat_qrcode_WeChatQRCode", None)
    if ctor is None:
        return []
    image_cv = cv2.imread(str(path))
    if image_cv is None:
        return []
    try:
        detector = ctor(*(str(p) for p in model_paths))
        decoded, points = detector.detectAndDecode(image_cv)
    except Exception:
        return []
    if not decoded:
        return []
    image = Image.open(path).convert("RGB")
    candidates: list[dict[str, Any]] = []
    for offset, payload in enumerate(decoded):
        payload = str(payload or "")
        if QR_PREPROCESS_REQUIRE_DECODE and not payload:
            continue
        if points is None or offset >= len(points):
            continue
        pts = [(float(x), float(y)) for x, y in points[offset]]
        bbox = _point_bbox(pts, image.size, padding=60)
        try:
            candidates.append(
                _qr_candidate_record(
                    engine="wechat",
                    source_path=path,
                    image=image,
                    bbox=bbox,
                    payload=payload,
                    output_dir=output_dir,
                    index=start_index + offset,
                )
            )
        except ValueError:
            continue
    image.close()
    return candidates


def _detect_qr_with_opencv(path: Path, output_dir: Path, *, start_index: int = 1) -> list[dict[str, Any]]:
    try:
        import cv2  # type: ignore
    except Exception:
        return []
    image_cv = cv2.imread(str(path))
    if image_cv is None:
        return []
    detector = cv2.QRCodeDetector()
    try:
        ok, decoded_info, points, _straight = detector.detectAndDecodeMulti(image_cv)
    except Exception:
        return []
    if not ok or points is None:
        return []
    image = Image.open(path).convert("RGB")
    candidates: list[dict[str, Any]] = []
    for offset, (pts, payload) in enumerate(zip(points, decoded_info)):
        payload = str(payload or "")
        if QR_PREPROCESS_REQUIRE_DECODE and not payload:
            continue
        xy = [(float(x), float(y)) for x, y in pts]
        bbox = _point_bbox(xy, image.size, padding=60)
        try:
            candidates.append(
                _qr_candidate_record(
                    engine="opencv",
                    source_path=path,
                    image=image,
                    bbox=bbox,
                    payload=payload,
                    output_dir=output_dir,
                    index=start_index + offset,
                )
            )
        except ValueError:
            continue
    image.close()
    return candidates


def extract_verified_qr_image(input_path: Path, output_dir: Path) -> dict[str, Any]:
    if not QR_PREPROCESS_ENABLED:
        return {"enabled": False, "status": "disabled"}
    output_dir.mkdir(parents=True, exist_ok=True)
    engines: list[tuple[str, Any]] = [
        ("zxing", _detect_qr_with_zxing),
        ("wechat", _detect_qr_with_wechat),
        ("opencv", _detect_qr_with_opencv),
    ]
    candidates: list[dict[str, Any]] = []
    seen_payloads: set[str] = set()
    engine_status: list[dict[str, Any]] = []
    for engine, fn in engines:
        try:
            found = fn(input_path, output_dir / engine, start_index=len(candidates) + 1)
        except Exception as exc:  # noqa: BLE001 - QR preprocess should be fail-soft
            engine_status.append({"engine": engine, "status": "error", "message": f"{type(exc).__name__}: {exc}"})
            continue
        accepted = []
        for item in found:
            payload_hash = str(item.get("payload_hash") or "")
            if payload_hash and payload_hash in seen_payloads:
                continue
            if payload_hash:
                seen_payloads.add(payload_hash)
            candidates.append(item)
            accepted.append(item)
        engine_status.append({"engine": engine, "status": "ok", "found": len(found), "accepted": len(accepted)})
    best = candidates[0] if candidates else None
    report = {
        "enabled": True,
        "status": "found" if best else "not_found",
        "source": str(input_path),
        "engines": engine_status,
        "candidate_count": len(candidates),
        "candidates": candidates,
        "image": best.get("image") if best else None,
        "payload_hash": best.get("payload_hash") if best else "",
        "message": "已提取可本地解码二维码。" if best else "未能从原始图片中稳定提取可本地解码二维码。",
    }
    report_path = output_dir / "qr_preprocess_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    report["report_file"] = str(report_path)
    return report


def detect_qr_crop_with_ocr_cli(input_path: Path, output_dir: Path) -> dict[str, Any]:
    if not QR_CROP_FALLBACK_ENABLED:
        return {"enabled": False, "status": "disabled"}
    if not OCR_CLI_PATH.is_file():
        return {"enabled": True, "status": "unavailable", "message": f"ocr-cli not found: {OCR_CLI_PATH}"}
    output_dir.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [
            str(OCR_CLI_PATH),
            "qr-detect",
            "--input",
            str(input_path),
            "--output-dir",
            str(output_dir),
        ],
        text=True,
        capture_output=True,
        timeout=60,
        check=False,
    )
    if proc.returncode != 0:
        return {
            "enabled": True,
            "status": "failed",
            "exit_code": proc.returncode,
            "stderr": proc.stderr[-800:],
        }
    try:
        payload = json.loads(proc.stdout.strip().splitlines()[-1])
    except (IndexError, json.JSONDecodeError) as exc:
        return {"enabled": True, "status": "invalid_output", "message": str(exc), "stdout": proc.stdout[-800:]}
    results = payload.get("results") or []
    first_image = next((item.get("image") for item in results if item.get("image")), None)
    return {
        "enabled": True,
        "status": "found" if first_image else "not_found",
        "count": int(payload.get("count") or len(results)),
        "result_file": payload.get("result_file"),
        "image": first_image,
        "decoded": next((item.get("decoded") for item in results if item.get("decoded")), ""),
    }


def load_single_summary(summary_path: Path) -> dict[str, Any]:
    if not summary_path.is_file():
        return {}
    return json.loads(summary_path.read_text(encoding="utf-8"))


def should_try_qr_crop_fallback(summary: dict[str, Any]) -> bool:
    text = " ".join(
        str(summary.get(key) or "")
        for key in ("business_result", "error_type", "stop_reason", "result_message")
    )
    return any(
        marker in text
        for marker in [
            "qr_unreadable",
            "未发现二维码",
            "未识别到二维码",
            "无法识别二维码",
            "album_entry_unavailable",
        ]
    )


def remote_stage_name(task_id: str, pass_name: str, attempt: int, source_path: str) -> str:
    suffix = Path(source_path).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".webp"}:
        suffix = ".png"
    stem = safe_filename_stem(f"invoice_{task_id}_{pass_name}_{attempt:02d}")
    return f"{stem}{suffix}"


def run_single_batch_attempt(
    args: argparse.Namespace,
    task: dict[str, Any],
    attempt_dir: Path,
    *,
    pass_name: str,
    attempt: int,
    stage_path: str,
) -> tuple[int, dict[str, Any]]:
    single_args = argparse.Namespace(**vars(args))
    single_args.output_dir = str(attempt_dir)
    single_args.stage_image = stage_path
    single_args.stage_remote_image = remote_stage_name(task["id"], pass_name, attempt, stage_path)
    single_args.stop_at = "result"
    single_args.return_to_entry = True
    try:
        code = run_agent(single_args)
        last_summary = load_single_summary(attempt_dir / "summary.json")
    except CancelledError:
        code = 130
        last_summary = {
            "status": "cancelled",
            "error_type": "cancelled_by_user",
            "stop_reason": "cancelled_by_user",
            "output_dir": str(attempt_dir),
        }
        attempt_dir.mkdir(parents=True, exist_ok=True)
        (attempt_dir / "summary.json").write_text(
            json.dumps(last_summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as exc:  # noqa: BLE001 - keep batch fail-soft on media/phone setup errors
        code = 2
        last_summary = {
            "status": "failed",
            "business_result": "technical_failed",
            "error_type": type(exc).__name__,
            "stop_reason": "phone_media_state_error",
            "result_message": str(exc),
            "output_dir": str(attempt_dir),
        }
        attempt_dir.mkdir(parents=True, exist_ok=True)
        (attempt_dir / "summary.json").write_text(
            json.dumps(last_summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return code, last_summary


def extract_result_conclusion(summary: dict[str, Any]) -> dict[str, Any]:
    validation_screenshot = summary.get("validation_screenshot")
    text = ""
    for step in summary.get("steps", []):
        if step.get("screenshot") == validation_screenshot:
            text = str(step.get("paddle_hint", ""))
            break
    if not text:
        for step in reversed(summary.get("steps", [])):
            decision = step.get("decision", {})
            if decision.get("result_status") == "success" and step.get("task_phase") == "verify":
                text = str(step.get("paddle_hint", ""))
                break
    classification = classify_business_result_text(text)
    business_result = summary.get("business_result") or classification.get("business_result")
    fields: dict[str, str | None] = {}
    if business_result == "verified" or (business_result is None and text):
        fields = {
            "票据代码": _extract_after_label(text, "票据代码"),
            "票据号码": _extract_after_label(text, "票据号码"),
            "校验码": _extract_after_label(text, "校验码"),
            "开票日期": _extract_after_label(text, "开票日期"),
            "金额合计": _extract_after_label(text, "金额合计"),
            "查验次数": _extract_after_label(text, "查验次数"),
        }
    return {
        "business_result": business_result,
        "result_message": summary.get("result_message") or classification.get("message"),
        "fields": {key: value for key, value in fields.items() if value},
        "ocr_excerpt": text[:1200],
    }


def _extract_after_label(text: str, label: str) -> str | None:
    pattern = re.compile(re.escape(label) + r"\s*[:：]?\s*([^\n\r|，,。 ]{1,64})")
    match = pattern.search(text)
    return match.group(1).strip() if match else None


def classify_business_result_text(text: str) -> dict[str, str]:
    normalized = _compact_text(text, limit=2000)
    compact = normalized.replace(" ", "")
    if not compact:
        return {}
    user_action_markers = [
        "锁屏",
        "输入密码",
        "登录微信",
        "请先登录",
        "权限申请",
        "权限设置",
        "验证码",
        "安全验证",
        "人脸识别",
        "身份验证",
    ]
    if any(marker in compact for marker in user_action_markers):
        return {
            "business_result": "user_action_required",
            "screen_state": "需要人工处理/权限/登录/安全验证",
            "message": normalized[:220],
            "stop_reason": "user_action_required",
        }
    daily_markers = [
        "超过当日限制查验次数",
        "超过当日查验次数",
        "当日限制查验次数",
        "请明日进行查验",
    ]
    if any(marker in compact for marker in daily_markers):
        return {
            "business_result": "daily_limit",
            "screen_state": "超过当日查验次数限制提示",
            "message": "超过当日可核验次数，请明日进行查验。",
            "stop_reason": "daily_limit_reached",
        }
    server_transient_markers = [
        "查验异常",
        "系统繁忙",
        "网络异常",
        "服务异常",
        "服务器异常",
        "请稍后再试",
        "暂时无法查验",
        "接口异常",
    ]
    if any(marker in compact for marker in server_transient_markers):
        return {
            "business_result": "server_transient",
            "screen_state": "服务端临时异常提示",
            "message": normalized[:220],
            "stop_reason": "server_transient_detected",
        }
    qr_unreadable_markers = [
        "未发现二维码",
        "未识别到二维码",
        "无法识别二维码",
        "不是有效的二维码",
        "未发现条码",
        "未发现小程序码",
        "识别失败",
    ]
    if any(marker in compact for marker in qr_unreadable_markers):
        return {
            "business_result": "qr_unreadable",
            "screen_state": "二维码不可识别提示",
            "message": normalized[:220],
            "stop_reason": "qr_unreadable_detected",
        }
    invalid_invoice_qr_markers = [
        "无效的电子票据二维码",
        "无效电子票据二维码",
        "无效二维码",
        "非财政票据二维码",
        "二维码格式不正确",
        "未识别到有效电子票据信息",
        "未识别有效电子票据信息",
    ]
    if any(marker in compact for marker in invalid_invoice_qr_markers):
        return {
            "business_result": "invalid_invoice_qr",
            "screen_state": "无效财政票据二维码提示",
            "message": normalized[:220],
            "stop_reason": "invalid_invoice_qr",
        }
    non_result_markers = [
        "相册",
        "图库",
        "选择图片",
        "扫码查验",
        "聊天输入",
        "发送",
        "服务号提供",
        "按住说话",
    ]
    has_result_core = "查验次数" in compact and "金额合计" in compact and ("票据代码" in compact or "票据号码" in compact)
    has_result_context = "财政票据" in compact or "票据详情" in compact or "票据信息" in compact
    if has_result_core and has_result_context and not any(marker in compact for marker in non_result_markers):
        return {
            "business_result": "verified",
            "screen_state": "票据详情/查验结果页",
            "message": "已获取票据详情页。",
            "stop_reason": "invoice_detail_captured",
        }
    invalid_markers = [
        "查无此票",
        "未查询到",
        "票据不存在",
        "二维码无效",
        "票据代码或号码错误",
        "票据信息有误",
        "查验失败",
        "不能查验",
        "无法查验",
    ]
    if any(marker in compact for marker in invalid_markers):
        return {
            "business_result": "invalid_or_not_found",
            "screen_state": "票据查验失败/不可查提示",
            "message": normalized[:220],
            "stop_reason": "invoice_invalid_or_not_found",
        }
    return {}


def write_batch_outputs(out_dir: Path, results: list[dict[str, Any]]) -> Path:
    screenshots_dir = out_dir / "validation_screenshots"
    screenshots_dir.mkdir(parents=True, exist_ok=True)
    for idx, item in enumerate(results, start=1):
        screenshot_path = item.get("validation_screenshot") or item.get("final_screenshot")
        if not screenshot_path:
            continue
        src = Path(screenshot_path)
        if src.is_file():
            screenshot_type = "核验结果截图" if item.get("validation_screenshot") else "失败现场截图"
            dst = screenshots_dir / validation_screenshot_filename(item, idx, src.suffix, screenshot_type)
            shutil.copy2(src, dst)
            item["packaged_validation_screenshot"] = str(dst)
            item["validation_screenshot_filename"] = dst.name
            item["screenshot_type"] = screenshot_type

    report_json = out_dir / "batch_report.json"
    report_json.write_text(json.dumps({"results": results}, ensure_ascii=False, indent=2), encoding="utf-8")

    daily_limit_manifest = out_dir / "daily_limit_retry_manifest.json"
    daily_limit_items = [
        {
            "task_id": item.get("task_id"),
            "source_name": item.get("source_name"),
            "qr_path": item.get("qr_path"),
            "business_result": item.get("business_result") or item.get("conclusion", {}).get("business_result"),
            "suggested_action": suggested_action_for_result(item),
            "reason": "当前票据达到当日查验次数上限",
        }
        for item in results
        if (item.get("business_result") or item.get("conclusion", {}).get("business_result")) == "daily_limit"
    ]
    if daily_limit_items:
        daily_limit_manifest.write_text(
            json.dumps({"items": daily_limit_items}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    report_xlsx = out_dir / "发票核验结果清单.xlsx"
    write_excel_report(report_xlsx, results)

    archive_path = out_dir / "invoice_check_results.zip"
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in [report_xlsx]:
            zf.write(path, path.relative_to(out_dir))
        if daily_limit_manifest.is_file():
            zf.write(daily_limit_manifest, daily_limit_manifest.relative_to(out_dir))
        for path in screenshots_dir.glob("*"):
            if path.is_file():
                zf.write(path, path.relative_to(out_dir))
    return archive_path


def validation_screenshot_filename(item: dict[str, Any], idx: int, suffix: str, screenshot_type: str = "核验结果截图") -> str:
    source = safe_filename_stem(str(item.get("source_name") or item.get("task_id") or f"invoice_{idx}"))
    type_stem = safe_filename_stem(screenshot_type)
    return f"{idx:03d}_{source}_{type_stem}{suffix or '.png'}"


def safe_filename_stem(value: str) -> str:
    stem = Path(value).stem or value
    stem = re.sub(r"[\\/:*?\"<>|]+", "_", stem)
    stem = re.sub(r"\s+", "_", stem).strip("._ ")
    return stem[:80] or "invoice"


def result_business_key(item: dict[str, Any]) -> str:
    return str(item.get("business_result") or item.get("conclusion", {}).get("business_result") or "")


def result_message_for_report(item: dict[str, Any]) -> str:
    message = item.get("conclusion", {}).get("result_message") or item.get("result_message") or ""
    if message == "No QR payload decoded by OpenCV.":
        return "未识别到有效二维码，请确认上传的是发票二维码截图。"
    business_result = result_business_key(item)
    error_type = str(item.get("error_type") or "")
    if business_result == "qr_unreadable_precheck":
        return "未识别到有效二维码，请确认上传的是发票二维码截图。"
    if business_result == "qr_extract_failed":
        return message or "未能从原始图片中稳定提取可本地解码二维码，请提供更清晰发票图片或人工核验。"
    if business_result == "invalid_invoice_qr":
        return message or "未识别到有效财政电子票据信息。"
    if business_result == "daily_limit":
        return message or "已达到当日查验次数限制，建议次日继续核验。"
    if business_result == "deferred_daily_limit":
        return message or "因当日查验次数限制，本张未查验，建议次日继续。"
    if business_result == "technical_navigation_failed" or error_type == "album_entry_unavailable":
        return message or "未能进入相册选择页，请稍后重试或人工复核。"
    if business_result == "qr_crop_retry_failed":
        return message or "整图扫码失败，QR 裁剪兜底仍未完成查验。"
    if business_result == "qr_not_found_after_correct_image_selected":
        return message or "已确认选择当前图片，但查验入口未识别到二维码。"
    if business_result == "field_fallback_failed":
        return message or "二维码与字段兜底均未完成查验，建议人工复核。"
    if error_type in {"same_state_loop", "cross_state_loop"}:
        return message or "手机页面多次重复未能完成操作，建议稍后重试或人工复核。"
    return str(message)


def suggested_action_for_result(item: dict[str, Any]) -> str:
    business_result = result_business_key(item)
    error_type = str(item.get("error_type") or "")
    if business_result == "verified":
        return "已完成，无需处理"
    if business_result == "daily_limit":
        return "次日继续核验"
    if business_result in {"deferred_daily_limit", "skipped_daily_limit"}:
        return "次日继续核验"
    if business_result == "qr_unreadable_precheck":
        return "重新上传清晰二维码或人工复核"
    if business_result == "qr_extract_failed":
        return "重新上传清晰发票图片或人工核验"
    if business_result == "invalid_invoice_qr":
        return "确认是否为财政电子票据二维码"
    if business_result == "server_transient":
        return "稍后重试"
    if business_result == "technical_navigation_failed" or error_type == "album_entry_unavailable":
        return "稍后重试或人工复核"
    if business_result in {"qr_crop_retry_failed", "qr_not_found_after_correct_image_selected", "field_fallback_failed"}:
        return "人工复核或重新上传清晰发票图片"
    if error_type in {"same_state_loop", "cross_state_loop", "max_steps_exceeded"}:
        return "技术失败重试或人工复核"
    if item.get("status") == "success":
        return "已完成，无需处理"
    return "人工复核"


def write_excel_report(path: Path, results: list[dict[str, Any]]) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "汇总"
    total = len(results)
    success = sum(1 for item in results if item.get("status") == "success")
    skipped = sum(1 for item in results if item.get("status") == "skipped")
    deferred = sum(1 for item in results if item.get("status") == "deferred")
    failed = total - success - skipped - deferred
    daily_limit_count = sum(1 for item in results if result_business_key(item) == "daily_limit")
    summary_rows = [
        ("校验总张数", total),
        ("已取得核验结果张数", success),
        ("因当日限制跳过张数", skipped),
        ("触发当日单票限额张数", daily_limit_count),
        ("延期/待后续核验张数", deferred),
        ("失败/需人工处理张数", failed),
        ("核验结果截图目录", "validation_screenshots/"),
        ("结果压缩包", "invoice_check_results.zip"),
    ]
    for row in summary_rows:
        summary_ws.append(row)
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 48
    for cell in summary_ws["A"]:
        cell.font = Font(bold=True)
    for row in summary_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws = wb.create_sheet("发票核验结果")
    ws.title = "发票核验结果"
    headers = [
        "序号",
        "来源文件",
        "任务状态",
        "核验结论",
        "结果说明",
        "尝试次数",
        "票据代码",
        "票据号码",
        "校验码",
        "开票日期",
        "金额合计",
        "查验次数",
        "错误类型",
        "建议动作",
        "截图类型",
        "核验结果截图文件名",
        "核验结果截图路径",
        "OCR摘要",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(results, start=1):
        fields = item.get("conclusion", {}).get("fields", {})
        conclusion = item.get("conclusion", {})
        packaged_shot = item.get("packaged_validation_screenshot") or item.get("validation_screenshot")
        packaged_name = item.get("validation_screenshot_filename") or (Path(packaged_shot).name if packaged_shot else "")
        row = [
            idx,
            item.get("source_name"),
            item.get("status"),
            conclusion.get("business_result") or item.get("business_result"),
            result_message_for_report(item),
            item.get("attempts"),
            fields.get("票据代码"),
            fields.get("票据号码"),
            fields.get("校验码"),
            fields.get("开票日期"),
            fields.get("金额合计"),
            fields.get("查验次数"),
            item.get("error_type"),
            suggested_action_for_result(item),
            item.get("screenshot_type") or ("核验结果截图" if item.get("validation_screenshot") else ("失败现场截图" if packaged_shot else "")),
            packaged_name,
            packaged_shot,
            item.get("conclusion", {}).get("ocr_excerpt"),
        ]
        ws.append(row)
        row_no = ws.max_row
        if packaged_shot:
            ws.cell(row=row_no, column=17).hyperlink = packaged_shot
            ws.cell(row=row_no, column=17).style = "Hyperlink"
        status_cell = ws.cell(row=row_no, column=3)
        if item.get("status") == "success":
            status_cell.fill = PatternFill("solid", fgColor="E2F0D9")
        elif item.get("status") == "deferred":
            status_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            status_cell.fill = PatternFill("solid", fgColor="FCE4D6")

    widths = [8, 32, 14, 22, 46, 10, 18, 18, 18, 16, 16, 16, 28, 24, 18, 36, 58, 80]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_single_result(args: argparse.Namespace, summary: dict[str, Any]) -> dict[str, Any]:
    source = Path(args.qr_image)
    return {
        "task_id": safe_task_id(source),
        "source_name": source.name,
        "qr_path": str(source.resolve()),
        "status": summary.get("status", "failed"),
        "attempts": 1,
        "attempt_records": [{
            "attempt": 1,
            "pass": "single",
            "summary_path": str(Path(args.output_dir).resolve() / "summary.json"),
            "status": summary.get("status", "failed"),
            "error_type": summary.get("error_type"),
        }],
        "summary_path": str(Path(args.output_dir).resolve() / "summary.json"),
        "validation_screenshot": summary.get("validation_screenshot"),
        "final_screenshot": summary.get("final_screenshot"),
        "stop_reason": summary.get("stop_reason"),
        "error_type": summary.get("error_type"),
        "business_result": summary.get("business_result"),
        "result_message": summary.get("result_message"),
        "conclusion": extract_result_conclusion(summary),
    }


def write_single_outputs(out_dir: Path, result: dict[str, Any]) -> Path:
    archive_path = write_batch_outputs(out_dir, [result])
    summary_path = out_dir / "single_summary.json"
    payload = {
        "status": result.get("status", "failed"),
        "total": 1,
        "success": 1 if result.get("status") == "success" else 0,
        "failed": 0 if result.get("status") == "success" else 1,
        "archive": str(archive_path),
        "report_json": str(out_dir / "batch_report.json"),
        "report_xlsx": str(out_dir / "发票核验结果清单.xlsx"),
        "screenshots_dir": str(out_dir / "validation_screenshots"),
        "result": result,
    }
    summary_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return archive_path


def archive_process_pages(out_dir: Path) -> Path | None:
    process_pages = sorted(out_dir.glob("step_*.png"))
    if not process_pages:
        return None
    archive_path = out_dir / "process_trace.zip"
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for page in process_pages:
            zf.write(page, page.name)
    for page in process_pages:
        try:
            page.unlink()
        except OSError:
            pass
    return archive_path


def archive_batch_process_pages(out_dir: Path) -> None:
    runs_dir = out_dir / "runs"
    if not runs_dir.is_dir():
        return
    for attempt_dir in runs_dir.glob("*/*/attempt_*"):
        if attempt_dir.is_dir():
            archive_process_pages(attempt_dir)


def emit_progress(message: str, **fields: Any) -> None:
    """Print a human-readable progress line for AgentD Task Output."""
    if not EMIT_PROGRESS_EVENTS:
        return
    suffix = ""
    if fields:
        clean_fields = {
            key: str(value)
            for key, value in fields.items()
            if value is not None and str(value) != ""
        }
        if clean_fields:
            suffix = " | " + " | ".join(f"{key}={value}" for key, value in clean_fields.items())
    print(f"Progress: {message}{suffix}", flush=True)


def _compact_text(value: Any, *, limit: int = 220) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > limit:
        return text[: limit - 1].rstrip() + "…"
    return text


def emit_agent(kind: str, message: str, **fields: Any) -> None:
    """Print a user-facing agent trace line for long-running Task Output."""
    labels = {
        "observe": "Agent观察",
        "think": "Agent判断",
        "act": "Agent动作",
        "result": "Agent结果",
    }
    label = labels.get(kind, "Agent日志")
    suffix = ""
    clean_fields = {
        key: _compact_text(value, limit=160)
        for key, value in fields.items()
        if value is not None and str(value) != ""
    }
    if clean_fields:
        suffix = "（" + "；".join(f"{key}：{value}" for key, value in clean_fields.items()) + "）"
    print(f"{label}：{message}{suffix}", flush=True)


def summarize_decision(decision: dict[str, Any], candidate: Candidate | None = None) -> str:
    action = str(decision.get("action") or "unknown")
    state = str(decision.get("screen_state") or "")
    reason = str(decision.get("reason") or "")
    target = ""
    if candidate is not None:
        x, y = candidate.center
        target = f" -> candidate {candidate.id} ({candidate.label_hint}) at ({x},{y})"
    summary = f"{action}{target}"
    if state:
        summary += f"; screen={state}"
    if reason:
        summary += f"; reason={reason}"
    return summary


def describe_action(decision: dict[str, Any], candidate: Candidate | None = None) -> str:
    action = str(decision.get("action") or "unknown").strip()
    if action == "tap_candidate" and candidate is not None:
        x, y = candidate.center
        return f"点击候选项 {candidate.id}（{candidate.label_hint}，坐标 {x},{y}）"
    mapping = {
        "wait": "等待页面响应",
        "home": "按 Android Home 键",
        "back": "按 Android Back 键",
        "swipe_left": "向左滑动桌面/页面",
        "swipe_right": "向右滑动桌面/页面",
        "stop_success": "停止并判定查验成功",
        "stop_failed": "停止并判定查验失败",
        "stop_user_action_required": "停止并等待用户人工处理",
    }
    return mapping.get(action, f"执行动作 {action}")


def canonicalize_action(action: str) -> str:
    normalized = str(action or "").strip().lower()
    aliases = {
        "click_candidate": "tap_candidate",
        "click": "tap_candidate",
        "tap": "tap_candidate",
        "press_candidate": "tap_candidate",
        "press": "tap_candidate",
    }
    return aliases.get(normalized, normalized)


def wants_search(decision: dict[str, Any]) -> bool:
    text = " ".join(str(decision.get(key, "")) for key in ("screen_state", "reason", "action")).lower()
    return any(marker in text for marker in ["search", "搜索", "查找"])


def is_wechat_search_page(decision: dict[str, Any], paddle_hint: str) -> bool:
    text = " ".join(
        str(value or "")
        for value in (
            decision.get("screen_state"),
            decision.get("reason"),
            paddle_hint,
        )
    ).lower()
    return (
        ("search page" in text or "搜索页" in text or "搜索页面" in text or "recent search" in text or "最近搜索" in text)
        and ("微信" in text or "wechat" in text)
    )


def wants_fiscal_ticket_candidate(decision: dict[str, Any], paddle_hint: str) -> bool:
    text = " ".join(
        str(value or "")
        for value in (
            decision.get("screen_state"),
            decision.get("reason"),
            paddle_hint,
        )
    ).lower()
    return "财政票据" in text and (
        "suggestion" in text
        or "recent" in text
        or "最近" in text
        or "候选" in text
        or "建议" in text
        or "搜索页" in text
        or "search page" in text
    )


def dangerous_for_search(label_hint: str) -> bool:
    return any(
        marker in label_hint
        for marker in [
            "top_left_close_or_back",
            "top_right_menu",
            "腾讯新闻",
            "cv_text_or_ui_region",
        ]
    )


def candidate_by_id(candidates: list[Candidate], candidate_id: Any) -> Candidate | None:
    try:
        cid = int(candidate_id or 0)
    except (TypeError, ValueError):
        return None
    return next((c for c in candidates if c.id == cid), None)


def find_candidate(candidates: list[Candidate], *markers: str) -> Candidate | None:
    for marker in markers:
        for candidate in candidates:
            if marker in candidate.label_hint:
                return candidate
    return None


def has_candidate(candidates: list[Candidate], marker: str) -> bool:
    return find_candidate(candidates, marker) is not None


def is_wechat_camera_scan_page_candidates(candidates: list[Candidate], paddle_hint: str = "") -> bool:
    compact = re.sub(r"\s+", "", paddle_hint or "").lower()
    scan_markers = ["我的二维码", "轻触照亮", "扫一扫", "扫码页", "扫描二维码", "scanqrcode"]
    return has_candidate(candidates, "camera_album_icon") and (
        has_candidate(candidates, "top_left_close_or_back")
        or any(marker in compact for marker in scan_markers)
    )


def looks_like_album_page(paddle_hint: str) -> bool:
    compact = re.sub(r"\s+", "", paddle_hint or "").lower()
    return any(
        marker in compact
        for marker in ["相册", "图库", "选择图片", "照片", "最近项目", "所有照片", "photopicker", "album", "gallery"]
    )


SAFE_CLEANUP_STATES = {
    "android_home",
    "wechat_home",
    "wechat_chat_list",
    "service_chat",
    "non_invoice_page",
}


def classify_cleanup_stable_screen(paddle_hint: str) -> str:
    compact = re.sub(r"\s+", "", paddle_hint or "").lower()
    raw_lower = (paddle_hint or "").lower()
    if not compact:
        return "unknown"

    business_popup_markers = [
        "超过当日限制查验次数",
        "超过当日查验次数",
        "查验异常",
        "无效的电子票据二维码",
        "无效电子票据二维码",
        "未发现二维码",
        "未识别到二维码",
        "确认",
        "确定",
    ]
    if any(marker in compact for marker in business_popup_markers):
        return "unsafe_business_popup"

    unsafe_markers = [
        "票据详情",
        "票据信息",
        "金额合计",
        "查验次数",
        "电子票号查验",
        "电子票据代码",
        "请输入电子票据代码",
        "扫码查验",
        "pjcy.mof.gov.cn",
        "相册",
        "图库",
        "选择图片",
        "photopicker",
        "album",
        "gallery",
        "我的二维码",
        "轻触照亮",
        "拍摄",
    ]
    if any(marker in compact or marker in raw_lower for marker in unsafe_markers):
        if any(marker in compact for marker in ["相册", "图库", "选择图片", "照片", "photopicker", "album", "gallery"]):
            return "unsafe_album_picker"
        if any(marker in compact for marker in ["我的二维码", "轻触照亮", "拍摄"]):
            return "unsafe_scan_camera"
        if any(marker in compact for marker in ["电子票号查验", "电子票据代码", "扫码查验", "pjcy.mof.gov.cn"]):
            return "unsafe_invoice_form"
        return "unsafe_invoice_detail"

    if looks_like_android_launcher(paddle_hint):
        return "android_home"

    wechat_nav_markers = ["微信", "搜索", "通讯录", "发现", "我"]
    if sum(1 for marker in wechat_nav_markers if marker in compact) >= 2:
        return "wechat_home"
    if any(marker in compact for marker in ["腾讯新闻", "订阅号", "服务通知"]) and "微信" in compact:
        return "wechat_chat_list"
    if "财政票据" in compact and any(marker in compact for marker in ["服务号", "公众号", "服务通知", "发消息", "聊天"]):
        return "service_chat"
    if "wechat" in raw_lower and not any(marker in compact for marker in ["票据", "扫码", "相册", "图库"]):
        return "wechat_home"
    return "unknown"


def wants_album_gallery(decision: dict[str, Any], paddle_hint: str = "") -> bool:
    text = " ".join(
        str(value or "")
        for value in (
            decision.get("screen_state"),
            decision.get("reason"),
            decision.get("action"),
            paddle_hint,
        )
    ).lower()
    return any(marker in text for marker in ["相册", "图库", "gallery", "album", "photo picker", "选择图片"])


def confirm_dialog_candidate(candidates: list[Candidate]) -> Candidate | None:
    return find_candidate(candidates, "dialog_confirm_button", "paddle_ocr:确认")


def h5_close_candidate(candidates: list[Candidate]) -> Candidate | None:
    return find_candidate(candidates, "top_left_close_or_back")


def close_dialog_if_present(candidates: list[Candidate], args: argparse.Namespace, *, reason: str) -> bool:
    confirm_candidate = confirm_dialog_candidate(candidates)
    if confirm_candidate is None:
        return False
    x, y = confirm_candidate.center
    emit_progress(
        "closing dialog after recording terminal state",
        reason=reason,
        candidate_id=confirm_candidate.id,
        center=f"{x},{y}",
    )
    emit_agent("act", "已先保存当前终态截图，现在关闭提示弹窗。", 原因=reason, 坐标=f"{x},{y}")
    tap(x, y)
    sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
    return True


def close_h5_context(candidates: list[Candidate], args: argparse.Namespace, *, reason: str) -> bool:
    close_candidate = h5_close_candidate(candidates)
    if close_candidate is not None:
        x, y = close_candidate.center
        emit_progress(
            "closing stale H5 context",
            reason=reason,
            candidate_id=close_candidate.id,
            center=f"{x},{y}",
        )
        emit_agent("act", "关闭当前财政票据 H5 上下文，下一步重新进入查验入口。", 原因=reason, 坐标=f"{x},{y}")
        tap(x, y)
        sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
        return True
    emit_progress("closing stale H5 context via Back", reason=reason)
    emit_agent("act", "未找到 H5 关闭按钮，使用返回键清理当前查验上下文。", 原因=reason)
    keyevent(4)
    sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
    return False


def reset_after_transient(candidates: list[Candidate], args: argparse.Namespace) -> None:
    closed = close_dialog_if_present(candidates, args, reason="server_transient")
    if closed:
        emit_agent("act", "服务端临时异常已关闭，继续关闭当前 H5 会话，下一次 attempt 将重新进入查验链路。")
    close_h5_context(candidates, args, reason="server_transient")


def cleanup_after_invoice_attempt(
    candidates: list[Candidate],
    args: argparse.Namespace,
    *,
    business_result: str,
    reason: str = "post_terminal",
    screen_text: str = "",
) -> dict[str, Any]:
    cleanup: dict[str, Any] = {
        "cleanup_attempted": True,
        "cleanup_status": "partial",
        "cleanup_action": "none",
    }
    manual_blocker_markers = [
        "locked",
        "锁屏",
        "permission",
        "权限",
        "captcha",
        "验证码",
        "security",
        "安全",
        "login",
        "登录",
        "credential",
        "凭证",
        "人脸",
    ]
    if business_result == "user_action_required" and any(marker in screen_text.lower() for marker in manual_blocker_markers):
        cleanup.update(
            {
                "cleanup_status": "skipped",
                "cleanup_warning": "manual_blocker_cleanup_skipped",
            }
        )
        return cleanup

    actions: list[str] = []
    try:
        if close_dialog_if_present(candidates, args, reason=f"{reason}:{business_result}"):
            actions.append("tap_confirm")

        if business_result == "server_transient":
            close_h5_context(candidates, args, reason=f"{reason}:{business_result}")
            actions.append("close_h5_or_back")
        elif not actions:
            emit_progress("post terminal cleanup via Back", reason=reason, business_result=business_result)
            emit_agent("act", "业务终态已记录，使用返回键清理当前手机页面，避免影响下一张票据。", 结论=business_result)
            keyevent(4)
            actions.append("back")
            sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))

        cleanup["cleanup_action"] = "+".join(actions) if actions else "none"
        cleanup["cleanup_status"] = "success" if actions else "partial"
        if not actions:
            cleanup["cleanup_warning"] = "no_cleanup_action_available"
        return cleanup
    except Exception as exc:  # noqa: BLE001 - cleanup must not overwrite business result
        cleanup.update(
            {
                "cleanup_status": "failed",
                "cleanup_action": "+".join(actions) if actions else "none",
                "cleanup_warning": f"{type(exc).__name__}: {exc}",
            }
        )
        return cleanup


def looks_like_h5_loading(text: str, decision: dict[str, Any] | None = None) -> bool:
    joined = " ".join(
        [
            str(text or ""),
            str((decision or {}).get("screen_state") or ""),
            str((decision or {}).get("reason") or ""),
        ]
    )
    raw_lower = joined.lower()
    compact = joined.replace(" ", "").lower()
    h5_markers = ["pjcy.mof.gov.cn", "财政票据", "票据查验", "h5", "loading", "加载"]
    form_markers = [
        "扫码查验",
        "电子票号查验",
        "电子票据代码",
        "请输入电子票据代码",
        "请输入票据号码",
        "票据号码输入",
        "票据代码输入",
    ]
    negative_form_markers = [
        "form controls not yet visible",
        "formcontrolsnotyetvisible",
        "not yet loaded",
        "notyetloaded",
        "not yet rendered",
        "notyetrendered",
        "no form controls",
        "noformcontrols",
        "form not visible",
        "formnotvisible",
        "scan tab not visible",
        "scantabnotvisible",
        "尚未显示表单",
        "未显示表单",
        "尚未显示扫码",
        "未显示扫码",
        "表单未加载",
    ]
    has_h5 = any(marker in compact for marker in h5_markers)
    has_loading = any(marker in compact for marker in ["loading", "加载", "空白", "headerpage", "bare", "blank"])
    has_negative_form = any(marker in raw_lower or marker in compact for marker in negative_form_markers)
    has_form = any(marker in compact for marker in form_markers)
    if has_negative_form and (has_h5 or has_loading):
        return True
    return has_h5 and has_loading and not has_form


def normalize_decision(
    decision: dict[str, Any],
    candidates: list[Candidate],
    paddle_hint: str,
    task_phase: str = "verify",
) -> dict[str, Any]:
    normalized = dict(decision)
    original_action = str(normalized.get("action") or "")
    action = canonicalize_action(original_action)
    if action != original_action:
        normalized["action"] = action
        normalized["host_guard"] = "action_canonicalized"
        normalized["original_action"] = original_action

    if task_phase == "verify" and is_wechat_camera_scan_page_candidates(candidates, paddle_hint):
        album_candidate = find_candidate(candidates, "camera_album_icon")
        if album_candidate:
            normalized["action"] = "tap_candidate"
            normalized["candidate_id"] = album_candidate.id
            normalized["screen_state"] = normalized.get("screen_state") or "wechat_camera_scan_page"
            normalized["result_status"] = "in_progress"
            normalized["host_guard"] = "camera_scan_album_hard_route"
            normalized["reason"] = (
                str(normalized.get("reason") or "")
                + "；机制层识别微信扫码页，强制点击右下角相册入口。"
            )
            action = "tap_candidate"
        else:
            normalized.update(
                {
                    "action": "stop_failed",
                    "candidate_id": None,
                    "screen_state": "wechat_camera_scan_page",
                    "result_status": "failed",
                    "business_result": "technical_navigation_failed",
                    "error_type": "album_entry_unavailable",
                    "host_guard": "camera_scan_album_candidate_missing",
                    "reason": "当前处于微信扫码页，但未生成相册入口候选，停止本票据以避免误点。",
                }
            )
            return normalized

    if action != "tap_candidate":
        return normalized

    selected = candidate_by_id(candidates, normalized.get("candidate_id"))
    if is_wechat_search_page(normalized, paddle_hint) and wants_fiscal_ticket_candidate(normalized, paddle_hint):
        fiscal_candidate = find_candidate(
            candidates,
            "paddle_ocr:财政票据",
            "wechat_recent_fiscal_ticket_search",
            "wechat_search_result_service_card",
            "wechat_search_result_ticket_check_button",
        )
        if fiscal_candidate and (selected is None or "wechat_home_search_icon" in selected.label_hint or dangerous_for_search(selected.label_hint)):
            normalized["candidate_id"] = fiscal_candidate.id
            normalized["host_guard"] = "candidate_reselected_fiscal_ticket"
            normalized["blocked_candidate"] = selected.as_json() if selected else None
            normalized["reason"] = (
                str(normalized.get("reason") or "")
                + "；机制层识别当前在微信搜索页，改选财政票据候选项。"
            )
            return normalized

    if wants_search(normalized):
        if is_wechat_search_page(normalized, paddle_hint):
            return normalized
        search_candidate = find_candidate(candidates, "wechat_home_search_icon", "wechat_search_input")
        if search_candidate and (selected is None or dangerous_for_search(selected.label_hint)):
            normalized["candidate_id"] = search_candidate.id
            normalized["host_guard"] = "candidate_reselected_search"
            normalized["blocked_candidate"] = selected.as_json() if selected else None
            normalized["reason"] = (
                str(normalized.get("reason") or "")
                + "；机制层将搜索意图重定向到确定的微信搜索候选。"
            )
            return normalized
        if selected is not None and dangerous_for_search(selected.label_hint):
            normalized.update(
                {
                    "action": "stop_failed",
                    "result_status": "failed",
                    "screen_state": normalized.get("screen_state") or "candidate_mismatch",
                    "error_type": "candidate_mismatch_search",
                    "host_guard": "candidate_mismatch_blocked",
                    "reason": "模型意图是搜索，但候选项不是搜索入口，已阻止误点。",
                }
            )
            return normalized

    if "launcher" in str(normalized.get("screen_state") or "").lower() and selected and "微信" not in selected.label_hint:
        wechat_candidate = find_candidate(candidates, "paddle_launcher_app:微信", "heuristic_launcher_app:微信", "paddle_ocr:微信")
        if wechat_candidate:
            normalized["candidate_id"] = wechat_candidate.id
            normalized["host_guard"] = "candidate_reselected_wechat_launcher"
            normalized["blocked_candidate"] = selected.as_json()
    return normalized


def loop_signature(paddle_hint: str, decision: dict[str, Any], candidate: Candidate | None) -> str:
    text_fingerprint = hashlib.sha1(_compact_text(paddle_hint, limit=500).encode("utf-8")).hexdigest()[:12]
    target = candidate.label_hint if candidate is not None else str(decision.get("candidate_id") or "")
    raw = "|".join(
        [
            text_fingerprint,
            str(decision.get("screen_state") or ""),
            str(decision.get("action") or ""),
            target,
        ]
    )
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def run_agent(args: argparse.Namespace) -> int:
    global EMIT_PROGRESS_EVENTS
    EMIT_PROGRESS_EVENTS = bool(getattr(args, "emit_progress_events", False))
    out_dir = Path(args.output_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    trace_path = out_dir / "trace.jsonl"
    summary_path = out_dir / "summary.json"

    emit_progress(
        "invoice verification started",
        output_dir=out_dir,
        max_steps=args.max_steps,
        stop_at=args.stop_at,
    )
    emit_agent(
        "act",
        "开始发票查验。我会先准备手机相册中的二维码，然后逐屏观察微信界面并决定下一步操作。",
        输出目录=out_dir,
        最大步数=args.max_steps,
    )
    width, height = get_size()
    emit_progress("phone screen detected", size=f"{width}x{height}")
    emit_agent("observe", "已连接到手机并读取屏幕尺寸。", 屏幕=f"{width}x{height}")
    summary: dict[str, Any] = {
        "status": "running",
        "stop_at": args.stop_at,
        "screen_size": [width, height],
        "steps": [],
        "output_dir": str(out_dir),
    }

    if args.stage_image:
        check_cancelled(getattr(args, "cancel_file", None))
        emit_progress("staging QR image to phone album", image=Path(args.stage_image).resolve(), remote_dir=REMOTE_DIR)
        emit_agent("act", "把待查验二维码推送到手机相册，供微信扫码页选择。", 文件=Path(args.stage_image).resolve())
        staged_remote = stage_image(
            Path(args.stage_image).resolve(),
            remote_name=str(getattr(args, "stage_remote_image", "") or REMOTE_IMAGE),
        )
        summary["staged_remote_image"] = staged_remote
        summary["phone_album_isolation"] = {"remote_dir": REMOTE_DIR, "remote_image": staged_remote, "file_count": 1}
        emit_progress("QR image staged", remote=staged_remote)
        emit_agent("result", "二维码已放入手机相册，且专用相册目录已隔离为当前单图。", 手机路径=staged_remote)
        sleep_with_cancel(args.after_stage_sleep, getattr(args, "cancel_file", None))

    server_transient_count = 0
    home_recover_count = 0
    loop_counts: dict[str, int] = {}
    transition_counts: dict[str, int] = {}
    transition_history: list[str] = []
    h5_loading_waits = 0
    album_entry_clicks = 0
    entered_phone_flow = bool(args.stage_image)
    has_selected_current_qr = False
    stale_terminal_cleanup_attempted = False
    last_candidates: list[Candidate] = []
    last_screen_text = ""
    task_phase = "verify"
    validation_screenshot: str | None = None
    for step in range(1, args.max_steps + 1):
        check_cancelled(getattr(args, "cancel_file", None))
        raw_path = out_dir / f"step_{step:03d}.png"
        annotated_path = out_dir / f"step_{step:03d}_candidates.png"
        emit_progress(f"step {step}/{args.max_steps}: capturing phone screen", screenshot=raw_path)
        emit_agent("observe", f"第 {step} 步：截取当前手机屏幕。")
        screenshot(raw_path)
        emit_progress(f"step {step}/{args.max_steps}: reading visible UI text")
        emit_agent("observe", f"第 {step} 步：识别屏幕文字和可点击区域。")
        paddle_hint = maybe_ask_paddle_vl(raw_path) if args.use_paddle_hint else ""
        paddle_boxes = paddle_text_boxes(paddle_hint, width, height) if args.use_paddle_hint else []
        if args.use_paddle_hint and not any("微信" in label for label, _box in paddle_boxes):
            paddle_boxes.extend(simplified_wechat_launcher_boxes(paddle_hint, width, height))
        page_is_launcher = looks_like_android_launcher(paddle_hint)
        candidates = build_candidates(raw_path, annotated_path, paddle_boxes, page_hint=paddle_hint)
        last_candidates = candidates
        last_screen_text = " ".join([paddle_hint, str(summary.get("error_type") or ""), str(summary.get("stop_reason") or "")])
        current_is_camera_scan_page = is_wechat_camera_scan_page_candidates(candidates, paddle_hint)
        current_is_album_page = looks_like_album_page(paddle_hint) and not current_is_camera_scan_page
        if current_is_camera_scan_page or current_is_album_page:
            entered_phone_flow = True
        if current_is_album_page:
            album_entry_clicks = 0
        elif not current_is_camera_scan_page:
            album_entry_clicks = 0
        cleanup_state = "unknown"
        if task_phase == "return_to_entry":
            cleanup_state = classify_cleanup_stable_screen(paddle_hint)
            if cleanup_state in SAFE_CLEANUP_STATES:
                summary["status"] = "success"
                summary["stop_reason"] = "returned_to_entry"
                summary["business_result"] = summary.get("business_result") or "verified"
                summary["result_message"] = summary.get("result_message") or "已查到票据详情。"
                summary["final_screenshot"] = str(raw_path)
                summary["cleanup_attempted"] = True
                summary["cleanup_status"] = "success"
                summary["cleanup_action"] = f"stable_stop:{cleanup_state}"
                if validation_screenshot:
                    summary["validation_screenshot"] = validation_screenshot
                emit_progress(
                    f"step {step}/{args.max_steps}: return cleanup reached stable screen",
                    cleanup_state=cleanup_state,
                    final_screenshot=raw_path,
                )
                emit_agent("result", "已离开发票查验流程并到达稳定页面，停止回退。", 稳定状态=cleanup_state)
                break
            if cleanup_state == "unsafe_business_popup":
                cleanup = cleanup_after_invoice_attempt(
                    candidates,
                    args,
                    business_result="return_to_entry_business_popup",
                    reason="return_to_entry_stable_guard",
                    screen_text=paddle_hint,
                )
                summary.setdefault("return_to_entry_cleanups", []).append(cleanup)
                continue
        emit_progress(
            f"step {step}/{args.max_steps}: candidates built",
            candidate_count=len(candidates),
            annotated=annotated_path,
        )
        hint_summary = _compact_text(paddle_hint, limit=180)
        if hint_summary and hint_summary.count("📶") < 5:
            emit_agent("observe", "屏幕理解结果已生成。", 画面=hint_summary)
        else:
            emit_agent("observe", "屏幕理解结果已生成。")
        emit_agent("observe", "已标注可点击候选区域。", 候选数量=len(candidates), 标注图=annotated_path)
        terminal_result = classify_business_result_text(paddle_hint)
        if (
            not terminal_result
            and task_phase == "verify"
            and has_selected_current_qr
            and looks_like_post_scan_return_to_service_page(paddle_hint)
        ):
            terminal_result = {
                "business_result": "qr_unreadable",
                "screen_state": "扫码后回到财政票据会话页",
                "message": "选择当前图片后未进入票据详情页，页面回到“轻触屏幕继续扫描”，按二维码未被识别处理。",
                "stop_reason": "post_scan_return_to_service_page",
            }
        if (
            terminal_result
            and task_phase == "verify"
            and args.stage_image
            and not has_selected_current_qr
            and not stale_terminal_cleanup_attempted
        ):
            stale_terminal_cleanup_attempted = True
            business_result = terminal_result["business_result"]
            emit_progress(
                f"step {step}/{args.max_steps}: stale terminal state before current QR selection",
                business_result=business_result,
                screenshot=raw_path,
            )
            emit_agent("act", "第一帧看到业务终态弹窗，但当前二维码尚未被选择，先按旧状态污染处理并清理。", 旧状态=business_result)
            cleanup = cleanup_after_invoice_attempt(
                candidates,
                args,
                business_result=business_result,
                reason="pre_current_qr_terminal",
                screen_text=paddle_hint,
            )
            cleanup["screenshot"] = str(raw_path)
            cleanup["business_result"] = business_result
            summary.setdefault("preflight_cleanups", []).append(cleanup)
            continue
        if terminal_result:
            business_result = terminal_result["business_result"]
            if business_result == "server_transient":
                result_status = "server_transient"
                action = "stop_failed"
            elif business_result == "user_action_required":
                result_status = "user_action_required"
                action = "stop_user_action_required"
            else:
                result_status = "success"
                action = "stop_success"
            decision = {
                "action": action,
                "candidate_id": None,
                "screen_state": terminal_result["screen_state"],
                "result_status": result_status,
                "business_result": business_result,
                "reason": terminal_result["message"],
                "confidence": 1.0,
            }
        else:
            album_candidate = find_candidate(candidates, "album_first_image_slot") if task_phase == "verify" else None
            if current_is_album_page and album_candidate is not None:
                decision = {
                    "action": "tap_candidate",
                    "candidate_id": album_candidate.id,
                    "screen_state": "photo_picker_album_page",
                    "result_status": "in_progress",
                    "reason": "机制层确认手机相册已隔离为当前单图，固定点击第一张图片。",
                    "confidence": 1.0,
                    "host_guard": "isolated_album_first_tile",
                }
            else:
                try:
                    decision = call_vlm(annotated_path, candidates, paddle_hint, step, args.stop_at, task_phase)
                except ModelDecisionError as exc:
                    decision = {
                        "action": "stop_failed",
                        "candidate_id": None,
                        "screen_state": "model_response_invalid",
                        "result_status": "failed",
                        "reason": str(exc),
                        "confidence": 1.0,
                        "error_type": "model_response_invalid",
                        "host_guard": "model_response_fail_soft",
                    }
            if "action" not in decision:
                decision = {
                    "action": "stop_failed",
                    "candidate_id": None,
                    "screen_state": "model_response_invalid",
                    "result_status": "failed",
                    "reason": "model decision missing action",
                    "confidence": 1.0,
                    "error_type": "model_response_invalid",
                    "host_guard": "model_response_fail_soft",
                }
            decision = normalize_decision(decision, candidates, paddle_hint, task_phase=task_phase)
        last_screen_text = " ".join(
            [
                paddle_hint,
                str(decision.get("screen_state") or ""),
                str(decision.get("reason") or ""),
                str(decision.get("result_status") or ""),
                str(decision.get("action") or ""),
            ]
        )
        planned_candidate = None
        if str(decision.get("action", "")).strip() == "tap_candidate":
            planned_candidate = candidate_by_id(candidates, decision.get("candidate_id"))
        emit_progress(
            f"step {step}/{args.max_steps}: decision",
            detail=summarize_decision(decision, planned_candidate),
        )
        emit_agent(
            "think",
            f"第 {step} 步决策完成。",
            当前画面=decision.get("screen_state"),
            理由=decision.get("reason"),
            计划动作=describe_action(decision, planned_candidate),
            置信度=decision.get("confidence"),
        )
        if not terminal_result and looks_like_h5_loading(paddle_hint, decision):
            h5_loading_waits += 1
        else:
            h5_loading_waits = 0

        record = {
            "step": step,
            "task_phase": task_phase,
            "screenshot": str(raw_path),
            "annotated": str(annotated_path),
            "candidate_count": len(candidates),
            "candidates": [c.as_json() for c in candidates],
            "paddle_hint": paddle_hint,
            "decision": decision,
            "h5_loading_waits": h5_loading_waits,
            "screen_flags": {
                "wechat_camera_scan_page": current_is_camera_scan_page,
                "album_page": current_is_album_page,
                "album_entry_clicks": album_entry_clicks,
                "cleanup_state": cleanup_state,
            },
        }
        signature = loop_signature(paddle_hint, decision, planned_candidate)
        record["loop_signature"] = signature
        loop_counts[signature] = loop_counts.get(signature, 0) + 1
        transition_token = "|".join(
            [
                str(decision.get("screen_state") or "")[:80],
                str(decision.get("action") or ""),
                planned_candidate.label_hint if planned_candidate else "",
            ]
        )
        transition_history.append(transition_token)
        if len(transition_history) > 8:
            transition_history = transition_history[-8:]
        transition_key = ">".join(transition_history[-4:])
        transition_counts[transition_key] = transition_counts.get(transition_key, 0) + 1
        record["transition_key"] = transition_key
        with trace_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
        summary["steps"].append(record)
        if args.emit_json_events:
            print(json.dumps({"step": step, "decision": decision}, ensure_ascii=False), flush=True)

        if (
            task_phase == "verify"
            and current_is_camera_scan_page
            and album_entry_clicks >= 2
            and find_candidate(candidates, "camera_album_icon")
        ):
            summary["status"] = "failed"
            summary["business_result"] = "technical_navigation_failed"
            summary["error_type"] = "album_entry_unavailable"
            summary["stop_reason"] = "album_entry_click_failed"
            summary["result_message"] = "未能进入相册选择页，请稍后重试或人工复核。"
            summary["final_screenshot"] = str(raw_path)
            emit_progress(
                f"step {step}/{args.max_steps}: album entry click failed",
                final_screenshot=raw_path,
                click_count=album_entry_clicks,
            )
            emit_agent("result", "相册入口连续点击后仍停留在扫码页，停止当前票据以避免误点。", 点击次数=album_entry_clicks)
            break

        if (
            not terminal_result
            and h5_loading_waits > int(getattr(args, "max_h5_loading_waits", MAX_H5_LOADING_WAITS))
        ):
            emit_progress(
                f"step {step}/{args.max_steps}: stale H5 loading detected",
                h5_loading_waits=h5_loading_waits,
                final_screenshot=raw_path,
            )
            emit_agent("act", "财政票据 H5 加载页连续停留过久，关闭当前 H5 后重新进入。", 次数=h5_loading_waits)
            close_h5_context(candidates, args, reason="h5_loading_stale")
            h5_loading_waits = 0
            continue

        if not terminal_result and loop_counts[signature] > int(getattr(args, "max_same_state_visits", MAX_SAME_STATE_VISITS)):
            summary["status"] = "failed"
            if current_is_camera_scan_page:
                summary["business_result"] = "technical_navigation_failed"
                summary["error_type"] = "album_entry_unavailable"
                summary["result_message"] = "未能进入相册选择页，请稍后重试或人工复核。"
            else:
                summary["error_type"] = "same_state_loop"
            summary["stop_reason"] = "same_state_loop"
            summary["final_screenshot"] = str(raw_path)
            emit_progress(
                f"step {step}/{args.max_steps}: same state loop detected",
                loop_signature=signature,
                count=loop_counts[signature],
                final_screenshot=raw_path,
            )
            emit_agent("result", "检测到同一画面/动作重复，停止当前票据以避免长循环。", 循环签名=signature)
            break

        if (
            not terminal_result
            and len(transition_history) >= 4
            and transition_history[-1] == transition_history[-3]
            and transition_history[-2] == transition_history[-4]
        ):
            summary["status"] = "failed"
            if current_is_camera_scan_page or "camera_album_icon" in transition_key:
                summary["business_result"] = "technical_navigation_failed"
                summary["error_type"] = "album_entry_unavailable"
                summary["result_message"] = "未能进入相册选择页，请稍后重试或人工复核。"
            else:
                summary["error_type"] = "cross_state_loop"
            summary["stop_reason"] = "cross_state_loop"
            summary["final_screenshot"] = str(raw_path)
            emit_progress(
                f"step {step}/{args.max_steps}: cross-state loop detected",
                transition_key=transition_key,
                final_screenshot=raw_path,
            )
            emit_agent("result", "检测到跨页面二元循环，停止当前票据以避免反复 Home/进入错误页面。", 循环=transition_key)
            break

        if terminal_result:
            business_result = terminal_result["business_result"]
            if business_result == "server_transient":
                summary["status"] = "failed"
                summary["error_type"] = "server_transient_detected"
            elif business_result == "user_action_required":
                summary["status"] = "user_action_required"
                summary["error_type"] = "user_action_required"
            else:
                summary["status"] = "success"
            summary["business_result"] = terminal_result["business_result"]
            summary["result_message"] = terminal_result["message"]
            summary["stop_reason"] = terminal_result["stop_reason"]
            summary["validation_screenshot"] = str(raw_path)
            summary["final_screenshot"] = str(raw_path)
            emit_agent(
                "result",
                "已识别到业务终态，停止当前票据查验，不再重复点击或重试。",
                结论=terminal_result["business_result"],
                说明=terminal_result["message"],
                结果截图=raw_path,
            )
            if business_result == "server_transient":
                summary["retry_recommended"] = True
            if args.return_to_entry:
                cleanup = cleanup_after_invoice_attempt(
                    candidates,
                    args,
                    business_result=business_result,
                    reason="terminal_result",
                    screen_text=paddle_hint,
                )
                summary.update(cleanup)
            break

        action = str(decision.get("action", "")).strip()
        result_status = str(decision.get("result_status", "")).strip()
        screen_state = str(decision.get("screen_state", "")).lower()
        if task_phase == "verify" and args.stop_at == "album" and (
            "photo picker" in screen_state
            or "album" in screen_state
            or "相册" in screen_state
            or "图库" in screen_state
        ):
            emit_progress(f"step {step}/{args.max_steps}: reached album page, stopping as requested")
            emit_agent("result", "已到达相册选择页，按当前 stop_at 设置停止。")
            summary["status"] = "success"
            summary["stop_reason"] = "stop_at_album"
            summary["final_screenshot"] = str(raw_path)
            break
        if action == "stop_success" or result_status == "success":
            if task_phase == "verify" and args.return_to_entry:
                validation_screenshot = str(raw_path)
                summary["validation_screenshot"] = validation_screenshot
                task_phase = "return_to_entry"
                emit_progress(
                    f"step {step}/{args.max_steps}: verification result captured, returning to entry page",
                    validation_screenshot=validation_screenshot,
                )
                emit_agent("result", "已经捕获查验结果页，接下来返回入口页。", 结果截图=validation_screenshot)
                if args.emit_json_events:
                    print(
                        json.dumps(
                            {
                                "step": step,
                                "host_guard": "validation_captured_return_to_entry",
                                "validation_screenshot": validation_screenshot,
                                "fallback": "back",
                            },
                            ensure_ascii=False,
                        ),
                        flush=True,
                    )
                keyevent(4)
                sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
                continue
            summary["status"] = "success"
            summary["final_screenshot"] = str(raw_path)
            if task_phase == "return_to_entry":
                summary["stop_reason"] = "returned_to_entry"
                summary["business_result"] = summary.get("business_result") or "verified"
                summary["result_message"] = summary.get("result_message") or "已查到票据详情。"
                summary["cleanup_attempted"] = summary.get("cleanup_attempted", True)
                summary["cleanup_status"] = summary.get("cleanup_status", "success")
                summary["cleanup_action"] = summary.get("cleanup_action", "return_to_entry")
                if validation_screenshot:
                    summary["validation_screenshot"] = validation_screenshot
            emit_progress(f"step {step}/{args.max_steps}: task succeeded", final_screenshot=raw_path)
            emit_agent("result", "查验流程成功结束。", 最终截图=raw_path)
            break
        if result_status == "server_transient":
            server_transient_count += 1
            emit_progress(
                f"step {step}/{args.max_steps}: server transient reported",
                count=server_transient_count,
                max_retries=args.max_server_retries,
            )
            emit_agent("think", "识别到服务端临时异常，先等待或重试，不立刻判定任务失败。", 次数=server_transient_count)
            if server_transient_count > args.max_server_retries:
                summary["status"] = "failed"
                summary["error_type"] = "max_server_transient_retries"
                summary["business_result"] = "server_transient"
                summary["retry_recommended"] = True
                summary["final_screenshot"] = str(raw_path)
                emit_progress("stopping after too many server transient errors", final_screenshot=raw_path)
                emit_agent("result", "连续服务端临时异常超过上限，停止本次查验。", 最终截图=raw_path)
                if args.return_to_entry:
                    cleanup = cleanup_after_invoice_attempt(
                        candidates,
                        args,
                        business_result="server_transient",
                        reason="max_server_transient_retries",
                        screen_text=paddle_hint,
                    )
                    summary.update(cleanup)
                break
        if action in {"stop_user_action_required", "stop_failed"}:
            if is_ignorable_system_banner(decision):
                emit_progress(f"step {step}/{args.max_steps}: ignorable system banner, waiting")
                emit_agent("think", "当前像是可忽略的系统提示，先等待页面恢复。")
                if args.emit_json_events:
                    print(
                        json.dumps(
                            {
                                "step": step,
                                "host_guard": "ignorable_system_banner_wait",
                                "fallback": "wait",
                            },
                            ensure_ascii=False,
                        ),
                        flush=True,
                    )
                sleep_with_cancel(args.wait_seconds, getattr(args, "cancel_file", None))
                continue
            if should_home_recover(decision) and home_recover_count < args.max_home_recovers:
                home_recover_count += 1
                emit_progress(
                    f"step {step}/{args.max_steps}: wrong app/page, pressing Home for recovery",
                    recover_count=home_recover_count,
                )
                emit_agent("act", "当前不在预期微信页面，按 Home 回到桌面后重新定位。", 恢复次数=home_recover_count)
                if args.emit_json_events:
                    print(
                        json.dumps(
                            {
                                "step": step,
                                "host_guard": "non_wechat_home_recover",
                                "home_recover_count": home_recover_count,
                            },
                            ensure_ascii=False,
                        ),
                        flush=True,
                    )
                keyevent(3)
                sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
                continue
            summary["status"] = "user_action_required" if action == "stop_user_action_required" else "failed"
            summary["error_type"] = decision.get("screen_state", action)
            summary["final_screenshot"] = str(raw_path)
            emit_progress(
                f"step {step}/{args.max_steps}: stopped",
                status=summary["status"],
                error_type=summary["error_type"],
                final_screenshot=raw_path,
            )
            emit_agent("result", "流程停止，需要用户查看原因或人工介入。", 状态=summary["status"], 原因=summary["error_type"])
            break
        if action == "wait":
            emit_progress(f"step {step}/{args.max_steps}: waiting", seconds=args.wait_seconds)
            emit_agent("act", "等待页面加载或状态变化。", 秒数=args.wait_seconds)
            sleep_with_cancel(args.wait_seconds, getattr(args, "cancel_file", None))
            continue
        if action == "home":
            emit_progress(f"step {step}/{args.max_steps}: pressing Android Home")
            emit_agent("act", "按 Android Home 键。")
            keyevent(3)
            sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
            continue
        if action == "back":
            emit_progress(f"step {step}/{args.max_steps}: pressing Android Back")
            emit_agent("act", "按 Android Back 键。")
            keyevent(4)
            sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
            continue
        if action == "swipe_left":
            if "launcher" in screen_state and ("not visible" in screen_state or "未" in str(decision.get("reason", ""))):
                emit_progress(f"step {step}/{args.max_steps}: launcher missing WeChat, pressing Home before swipe")
                emit_agent("act", "桌面上没有看到微信，先按 Home 复位再继续。")
                if args.emit_json_events:
                    print(
                        json.dumps(
                            {"step": step, "host_guard": "launcher_missing_wechat_home_first", "fallback": "home"},
                            ensure_ascii=False,
                        ),
                        flush=True,
                    )
                keyevent(3)
                sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
                continue
            emit_progress(f"step {step}/{args.max_steps}: swiping left")
            emit_agent("act", "向左滑动，尝试寻找微信或目标入口。")
            swipe(int(width * 0.82), int(height * 0.55), int(width * 0.18), int(height * 0.55))
            sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
            continue
        if action == "swipe_right":
            if "launcher" in screen_state and ("not visible" in screen_state or "未" in str(decision.get("reason", ""))):
                emit_progress(f"step {step}/{args.max_steps}: launcher missing WeChat, pressing Home before swipe")
                emit_agent("act", "桌面上没有看到微信，先按 Home 复位再继续。")
                if args.emit_json_events:
                    print(
                        json.dumps(
                            {"step": step, "host_guard": "launcher_missing_wechat_home_first", "fallback": "home"},
                            ensure_ascii=False,
                        ),
                        flush=True,
                    )
                keyevent(3)
                sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
                continue
            emit_progress(f"step {step}/{args.max_steps}: swiping right")
            emit_agent("act", "向右滑动，尝试寻找微信或目标入口。")
            swipe(int(width * 0.18), int(height * 0.55), int(width * 0.82), int(height * 0.55))
            sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
            continue
        if action == "tap_candidate":
            cid = int(decision.get("candidate_id") or 0)
            candidate = next((c for c in candidates if c.id == cid), None)
            if candidate is None:
                if page_is_launcher:
                    emit_progress(
                        f"step {step}/{args.max_steps}: invalid launcher candidate, swiping left instead",
                        bad_candidate_id=cid,
                    )
                    emit_agent("think", "模型给出的桌面候选项无效，改为滑动查找。", 候选编号=cid)
                    if args.emit_json_events:
                        print(
                            json.dumps(
                                {
                                    "step": step,
                                    "host_guard": "launcher_invalid_candidate_blocked",
                                    "bad_candidate_id": cid,
                                    "fallback": "swipe_left",
                                },
                                ensure_ascii=False,
                            ),
                            flush=True,
                    )
                    swipe(int(width * 0.82), int(height * 0.55), int(width * 0.18), int(height * 0.55))
                    sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
                    continue
                summary["status"] = "failed"
                summary["error_type"] = "invalid_candidate_id"
                summary["bad_decision"] = decision
                summary["final_screenshot"] = str(raw_path)
                emit_progress(
                    f"step {step}/{args.max_steps}: invalid candidate, stopping",
                    bad_candidate_id=cid,
                    final_screenshot=raw_path,
                )
                emit_agent("result", "模型给出的候选编号不存在，停止以避免误点。", 候选编号=cid)
                break
            if page_is_launcher and "微信" not in candidate.label_hint:
                emit_progress(
                    f"step {step}/{args.max_steps}: blocked non-WeChat launcher tap, swiping left instead",
                    candidate=candidate.label_hint,
                )
                emit_agent("think", "候选项不是微信入口，阻止误点并改为继续滑动。", 候选=candidate.label_hint)
                if args.emit_json_events:
                    print(
                        json.dumps(
                            {
                                "step": step,
                                "host_guard": "launcher_non_wechat_tap_blocked",
                                "blocked_candidate": candidate.as_json(),
                                "fallback": "swipe_left",
                            },
                            ensure_ascii=False,
                        ),
                        flush=True,
                )
                swipe(int(width * 0.82), int(height * 0.55), int(width * 0.18), int(height * 0.55))
                sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
                continue
            x, y = candidate.center
            emit_progress(
                f"step {step}/{args.max_steps}: tapping candidate",
                candidate_id=candidate.id,
                label=candidate.label_hint,
                center=f"{x},{y}",
            )
            emit_agent("act", f"点击候选项 {candidate.id}。", 名称=candidate.label_hint, 坐标=f"{x},{y}")
            tap(x, y)
            if candidate.label_hint == "camera_album_icon":
                album_entry_clicks += 1
                entered_phone_flow = True
            else:
                album_entry_clicks = 0
            decision_screen_state = str(decision.get("screen_state") or "").lower()
            if (
                (current_is_album_page or "photo_picker" in decision_screen_state or "album" in decision_screen_state)
                and "top_left_close_or_back" not in candidate.label_hint
            ):
                has_selected_current_qr = True
                entered_phone_flow = True
            if candidate.label_hint == "album_first_image_slot":
                has_selected_current_qr = True
                entered_phone_flow = True
            sleep_with_cancel(args.after_action_sleep, getattr(args, "cancel_file", None))
            continue

        summary["status"] = "failed"
        summary["error_type"] = "unknown_action"
        summary["bad_decision"] = decision
        summary["final_screenshot"] = str(raw_path)
        emit_progress(
            f"step {step}/{args.max_steps}: unknown action, stopping",
            action=action,
            final_screenshot=raw_path,
        )
        emit_agent("result", "收到未知动作，停止以避免误操作。", 动作=action)
        break
    else:
        if validation_screenshot:
            summary["status"] = "success"
            summary["business_result"] = summary.get("business_result") or "verified"
            summary["result_message"] = summary.get("result_message") or "已查到票据详情。"
            summary["stop_reason"] = "success_with_return_warning"
            summary["cleanup_warning"] = "return_to_entry_max_steps_exceeded"
            summary["validation_screenshot"] = validation_screenshot
            emit_progress("maximum step count reached after validation screenshot", max_steps=args.max_steps)
            emit_agent("result", "已取得查验结果截图，但返回入口阶段超过步数；保留票据核验成功结果。", 最大步数=args.max_steps)
        else:
            summary["status"] = "failed"
            summary["error_type"] = "max_steps_exceeded"
            emit_progress("maximum step count reached", max_steps=args.max_steps)
            emit_agent("result", "达到最大步数，本次短跑停止。", 最大步数=args.max_steps)

    if entered_phone_flow and args.return_to_entry and not summary.get("cleanup_attempted"):
        cleanup_business_result = str(
            summary.get("business_result")
            or summary.get("error_type")
            or summary.get("stop_reason")
            or "technical_navigation_failed"
        )
        cleanup = cleanup_after_invoice_attempt(
            last_candidates,
            args,
            business_result=cleanup_business_result,
            reason="finalize_single_attempt",
            screen_text=last_screen_text,
        )
        summary.update(cleanup)

    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    emit_progress("invoice verification finished", status=summary["status"], summary=summary_path)
    emit_agent("result", "本次发票查验任务已写入结果文件。", 状态=summary["status"], 摘要=summary_path)
    if args.emit_json_events:
        print(json.dumps({"summary": str(summary_path), "status": summary["status"]}, ensure_ascii=False))
    return 0 if summary["status"] == "success" else 2


def run_batch(args: argparse.Namespace) -> int:
    out_dir = Path(args.output_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    input_dir = Path(args.batch_input_dir).resolve()
    tasks = discover_qr_tasks(input_dir)
    if not tasks:
        raise SystemExit(f"No QR image files found in {input_dir}")

    results_by_id: dict[str, dict[str, Any]] = {}
    failed: list[dict[str, Any]] = []
    for index, task in enumerate(tasks):
        check_cancelled(getattr(args, "cancel_file", None))
        result = run_batch_task(args, task, out_dir, pass_name="first")
        results_by_id[task["id"]] = result
        if result["status"] != "success":
            failed.append(task)
            if args.failed_task_sleep > 0:
                sleep_with_cancel(args.failed_task_sleep, getattr(args, "cancel_file", None))

    if failed and not args.no_failed_replay:
        replay_failed: list[dict[str, Any]] = []
        if args.failed_replay_sleep > 0:
            sleep_with_cancel(args.failed_replay_sleep, getattr(args, "cancel_file", None))
        for task in failed:
            check_cancelled(getattr(args, "cancel_file", None))
            result = run_batch_task(args, task, out_dir, pass_name="replay")
            results_by_id[task["id"]] = result
            if result["status"] != "success":
                replay_failed.append(task)
                if args.failed_task_sleep > 0:
                    sleep_with_cancel(args.failed_task_sleep, getattr(args, "cancel_file", None))
        failed = replay_failed

    ordered_results = [results_by_id[task["id"]] for task in tasks]
    archive_path = write_batch_outputs(out_dir, ordered_results)
    archive_batch_process_pages(out_dir)
    deferred = sum(1 for item in ordered_results if item["status"] == "deferred")
    technical_failed = sum(1 for item in ordered_results if item["status"] not in {"success", "skipped", "deferred"})
    daily_limit_count = sum(1 for item in ordered_results if result_business_key(item) == "daily_limit")
    business_result_counts: dict[str, int] = {}
    for item in ordered_results:
        key = result_business_key(item) or str(item.get("error_type") or item.get("status") or "unknown")
        business_result_counts[key] = business_result_counts.get(key, 0) + 1
    batch_summary = {
        "status": "success" if not failed else "partial_failed",
        "total": len(tasks),
        "success": sum(1 for item in ordered_results if item["status"] == "success"),
        "skipped": sum(1 for item in ordered_results if item["status"] == "skipped"),
        "deferred": deferred,
        "daily_limit": daily_limit_count,
        "business_result_counts": business_result_counts,
        "technical_failed": technical_failed,
        "failed": technical_failed,
        "archive": str(archive_path),
        "report_json": str(out_dir / "batch_report.json"),
        "report_xlsx": str(out_dir / "发票核验结果清单.xlsx"),
        "screenshots_dir": str(out_dir / "validation_screenshots"),
        "results": ordered_results,
    }
    (out_dir / "batch_summary.json").write_text(
        json.dumps(batch_summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(batch_summary, ensure_ascii=False))
    # Mixed business outcomes are still a successful batch delivery when the
    # Excel/zip report is produced and no technical item failed. Keep terminal
    # invoice statuses in the report instead of making AgentD mark the detached
    # process as failed.
    return 0 if technical_failed == 0 else 2


def run_batch_task(args: argparse.Namespace, task: dict[str, Any], batch_out_dir: Path, *, pass_name: str) -> dict[str, Any]:
    attempts_dir = batch_out_dir / "runs" / task["id"] / pass_name
    attempts_dir.mkdir(parents=True, exist_ok=True)
    last_summary: dict[str, Any] = {}
    attempt_records: list[dict[str, Any]] = []
    qr_preprocess = extract_verified_qr_image(Path(task["qr_path"]), attempts_dir / "qr_preprocess")
    if QR_PREPROCESS_ENABLED and qr_preprocess.get("status") != "found":
        result_message = str(
            qr_preprocess.get("message")
            or "未能从原始图片中稳定提取可本地解码二维码，请用户提供更清晰图片或人工核验。"
        )
        return {
            "task_id": task["id"],
            "source_name": task["source_name"],
            "qr_path": task["qr_path"],
            "status": "deferred",
            "attempts": 0,
            "attempt_records": [],
            "summary_path": None,
            "validation_screenshot": None,
            "final_screenshot": None,
            "stop_reason": "qr_extract_failed",
            "error_type": None,
            "business_result": "qr_extract_failed",
            "result_message": result_message,
            "qr_preprocess": qr_preprocess,
            "qr_precheck": {"enabled": QR_PRECHECK_ENABLED, "status": "skipped_after_preprocess_failed"},
            "qr_crop_fallback": {"enabled": QR_CROP_FALLBACK_ENABLED, "status": "skipped_after_preprocess_failed"},
            "conclusion": {
                "business_result": "qr_extract_failed",
                "result_message": result_message,
                "fields": {},
                "ocr_excerpt": "",
            },
        }
    stage_input_path = str(qr_preprocess.get("image") or task["qr_path"])
    qr_precheck = precheck_qr_image(Path(task["qr_path"]))
    if (
        QR_PRECHECK_MODE == "block_unreadable"
        and qr_precheck.get("status") in {"missing", "unreadable"}
    ):
        result_message = str(qr_precheck.get("message") or "二维码图片预检未通过。")
        return {
            "task_id": task["id"],
            "source_name": task["source_name"],
            "qr_path": task["qr_path"],
            "status": "success",
            "attempts": 0,
            "attempt_records": [],
            "summary_path": None,
            "validation_screenshot": None,
            "final_screenshot": None,
            "stop_reason": "qr_unreadable_precheck",
            "error_type": None,
            "business_result": "qr_unreadable_precheck",
            "result_message": result_message,
            "qr_precheck": qr_precheck,
            "conclusion": {
                "business_result": "qr_unreadable_precheck",
                "result_message": result_message,
                "fields": {},
                "ocr_excerpt": "",
            },
        }
    qr_crop_fallback = (
        detect_qr_crop_with_ocr_cli(Path(task["qr_path"]), attempts_dir / "qr_crop_fallback")
        if (not qr_preprocess.get("image")) and qr_precheck.get("status") == "unreadable"
        else {"enabled": QR_CROP_FALLBACK_ENABLED, "status": "not_needed"}
    )
    for attempt in range(1, args.batch_attempts + 1):
        check_cancelled(getattr(args, "cancel_file", None))
        attempt_dir = attempts_dir / f"attempt_{attempt:02d}"
        code, last_summary = run_single_batch_attempt(
            args,
            task,
            attempt_dir,
            pass_name=pass_name,
            attempt=attempt,
            stage_path=stage_input_path,
        )
        if not args.keep_phone_album and code != 130 and not is_cancelled(getattr(args, "cancel_file", None)):
            cleanup_phone_album()
        attempt_records.append({
            "attempt": attempt,
            "pass": pass_name,
            "exit_code": code,
            "summary_path": str(attempt_dir / "summary.json"),
            "status": last_summary.get("status", "failed"),
            "error_type": last_summary.get("error_type"),
            "qr_preprocess": qr_preprocess,
            "qr_precheck": qr_precheck,
            "qr_crop_fallback": qr_crop_fallback,
            "stage_path": stage_input_path,
        })
        if code == 0 and last_summary.get("status") == "success":
            conclusion = extract_result_conclusion(last_summary)
            return {
                "task_id": task["id"],
                "source_name": task["source_name"],
                "qr_path": task["qr_path"],
                "status": "success",
                "attempts": len(attempt_records),
                "attempt_records": attempt_records,
                "summary_path": str(attempt_dir / "summary.json"),
                "validation_screenshot": last_summary.get("validation_screenshot"),
                "final_screenshot": last_summary.get("final_screenshot"),
                "stop_reason": last_summary.get("stop_reason"),
                "business_result": last_summary.get("business_result"),
                "result_message": last_summary.get("result_message"),
                "qr_preprocess": qr_preprocess,
                "qr_precheck": qr_precheck,
                "qr_crop_fallback": qr_crop_fallback,
                "conclusion": conclusion,
            }
        if code == 130 or last_summary.get("status") == "cancelled":
            break
        if args.attempt_sleep > 0:
            sleep_with_cancel(args.attempt_sleep, getattr(args, "cancel_file", None))

    fallback_image = qr_crop_fallback.get("image")
    if fallback_image and should_try_qr_crop_fallback(last_summary):
        attempt = len(attempt_records) + 1
        attempt_dir = attempts_dir / f"attempt_{attempt:02d}_qr_crop"
        code, last_summary = run_single_batch_attempt(
            args,
            task,
            attempt_dir,
            pass_name=f"{pass_name}_qr_crop",
            attempt=attempt,
            stage_path=str(fallback_image),
        )
        if not args.keep_phone_album and code != 130 and not is_cancelled(getattr(args, "cancel_file", None)):
            cleanup_phone_album()
        if code != 0 and last_summary.get("business_result") in {"qr_unreadable", None}:
            last_summary["business_result"] = "qr_crop_retry_failed"
            last_summary["result_message"] = last_summary.get("result_message") or "整图扫码失败，QR 裁剪兜底仍未完成查验。"
            (attempt_dir / "summary.json").write_text(json.dumps(last_summary, ensure_ascii=False, indent=2), encoding="utf-8")
        attempt_records.append({
            "attempt": attempt,
            "pass": f"{pass_name}_qr_crop",
            "exit_code": code,
            "summary_path": str(attempt_dir / "summary.json"),
            "status": last_summary.get("status", "failed"),
            "error_type": last_summary.get("error_type"),
            "qr_precheck": qr_precheck,
            "qr_crop_fallback": qr_crop_fallback,
            "stage_path": str(fallback_image),
        })
        if code == 0 and last_summary.get("status") == "success":
            conclusion = extract_result_conclusion(last_summary)
            return {
                "task_id": task["id"],
                "source_name": task["source_name"],
                "qr_path": task["qr_path"],
                "status": "success",
                "attempts": len(attempt_records),
                "attempt_records": attempt_records,
                "summary_path": str(attempt_dir / "summary.json"),
                "validation_screenshot": last_summary.get("validation_screenshot"),
                "final_screenshot": last_summary.get("final_screenshot"),
                "stop_reason": last_summary.get("stop_reason"),
                "business_result": last_summary.get("business_result"),
                "result_message": last_summary.get("result_message"),
                "qr_precheck": qr_precheck,
                "qr_crop_fallback": qr_crop_fallback,
                "conclusion": conclusion,
            }

    return {
        "task_id": task["id"],
        "source_name": task["source_name"],
        "qr_path": task["qr_path"],
        "status": last_summary.get("status", "failed") if last_summary else "failed",
        "attempts": len(attempt_records),
        "attempt_records": attempt_records,
        "summary_path": attempt_records[-1]["summary_path"] if attempt_records else None,
        "validation_screenshot": last_summary.get("validation_screenshot") if last_summary else None,
        "final_screenshot": last_summary.get("final_screenshot") if last_summary else None,
        "error_type": last_summary.get("error_type") if last_summary else "no_summary",
        "business_result": last_summary.get("business_result") if last_summary else None,
        "result_message": last_summary.get("result_message") if last_summary else None,
        "qr_preprocess": qr_preprocess,
        "qr_precheck": qr_precheck,
        "qr_crop_fallback": qr_crop_fallback,
        "conclusion": extract_result_conclusion(last_summary) if last_summary else {},
    }


def cmd_health(_: argparse.Namespace) -> int:
    payload = {
        "success": True,
        "service": "invoice-check-cli",
        "config": {
            "adb_path": str(ADB),
            "vlm_api": VLM_API,
            "vlm_model": VLM_MODEL,
            "paddle_vl_api": PADDLE_VL_API,
            "paddle_vl_model": PADDLE_VL_MODEL,
            "remote_dir": REMOTE_DIR,
            "remote_image": REMOTE_IMAGE,
            "adb_serial": ADB_SERIAL,
            "runtime_dir": str(RUNTIME_DIR),
            "queue_db_path": str(QUEUE_DB_PATH),
            "phone_resource": {
                "resource_id": PHONE_RESOURCE_ID,
                "display_name": PHONE_DISPLAY_NAME,
                "lock_wait_seconds": PHONE_LOCK_WAIT_SECONDS,
                "queue_wait_timeout_seconds": QUEUE_WAIT_TIMEOUT_SECONDS,
            },
            "runtime_gates": {
                "max_steps": int(CONFIG.get("max_steps", 30)),
                "batch_attempts": int(CONFIG.get("batch_attempts", 1)),
                "failed_replay_enabled": bool(CONFIG.get("failed_replay_enabled", False)),
                "cleanup_gallery_for_invoice": CLEANUP_GALLERY_FOR_INVOICE,
                "max_same_state_visits": MAX_SAME_STATE_VISITS,
                "max_h5_loading_waits": MAX_H5_LOADING_WAITS,
                "qr_precheck_enabled": QR_PRECHECK_ENABLED,
                "qr_precheck_block_unreadable": QR_PRECHECK_BLOCK_UNREADABLE,
                "qr_precheck_mode": QR_PRECHECK_MODE,
                "qr_preprocess_enabled": QR_PREPROCESS_ENABLED,
                "qr_preprocess_require_decode": QR_PREPROCESS_REQUIRE_DECODE,
                "qr_preprocess_canvas_size": QR_PREPROCESS_CANVAS_SIZE,
                "qr_preprocess_qr_max_size": QR_PREPROCESS_QR_MAX_SIZE,
                "qr_wechat_model_dir": str(QR_WECHAT_MODEL_DIR),
                "qr_wechat_models_available": _wechat_model_paths() is not None,
                "qr_crop_fallback_enabled": QR_CROP_FALLBACK_ENABLED,
                "ocr_cli_path": str(OCR_CLI_PATH),
            },
            "vlm_sampling": {
                "temperature": VLM_TEMPERATURE,
                "top_p": VLM_TOP_P,
                "top_k": VLM_TOP_K,
                "max_tokens": VLM_MAX_TOKENS,
                "timeout": VLM_TIMEOUT,
                "enable_thinking": VLM_ENABLE_THINKING,
            },
            "paddle_vl_sampling": {
                "temperature": PADDLE_VL_TEMPERATURE,
                "max_tokens": PADDLE_VL_MAX_TOKENS,
                "timeout": PADDLE_VL_TIMEOUT,
            },
        },
    }
    print(json.dumps(payload, ensure_ascii=False))
    return 0


def add_runtime_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--max-steps", type=int, default=int(CONFIG.get("max_steps", 80)))
    parser.add_argument("--max-server-retries", type=int, default=int(CONFIG.get("max_server_retries", 2)))
    parser.add_argument("--max-same-state-visits", type=int, default=MAX_SAME_STATE_VISITS)
    parser.add_argument("--max-same-action-repeats", type=int, default=MAX_SAME_ACTION_REPEATS)
    parser.add_argument("--max-h5-loading-waits", type=int, default=MAX_H5_LOADING_WAITS)
    parser.add_argument("--after-stage-sleep", type=float, default=float(CONFIG.get("after_stage_sleep", 1.5)))
    parser.add_argument("--after-action-sleep", type=float, default=float(CONFIG.get("after_action_sleep", 2.5)))
    parser.add_argument("--wait-seconds", type=float, default=float(CONFIG.get("wait_seconds", 3.0)))
    parser.add_argument("--max-home-recovers", type=int, default=int(CONFIG.get("max_home_recovers", 3)))
    parser.add_argument("--use-paddle-hint", action=argparse.BooleanOptionalAction, default=bool(CONFIG.get("use_paddle_hint", True)))
    parser.add_argument("--cleanup-phone-album", action=argparse.BooleanOptionalAction, default=bool(CONFIG.get("cleanup_phone_album", True)))
    parser.add_argument("--keep-phone-album", action="store_true")
    parser.add_argument("--cancel-file", default=os.environ.get("INVOICE_CHECK_CANCEL_FILE") or os.environ.get("AGENTD_CANCEL_FILE"))
    parser.add_argument("--phone-lock-wait-seconds", type=float, default=PHONE_LOCK_WAIT_SECONDS)
    parser.add_argument(
        "--emit-json-events",
        action="store_true",
        help="Also print per-step JSON debug events to stdout. Human-readable Task Output is the default.",
    )
    parser.add_argument(
        "--emit-progress-events",
        action="store_true",
        help="Also print low-level Progress: debug lines to stdout. Hidden by default.",
    )


def add_batch_policy_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--batch-attempts", "--attempts", type=int, default=int(CONFIG.get("batch_attempts", 3)))
    parser.add_argument("--attempt-sleep", type=float, default=float(CONFIG.get("attempt_sleep", 2.0)))
    parser.add_argument("--failed-task-sleep", type=float, default=float(CONFIG.get("failed_task_sleep", 10.0)))
    parser.add_argument("--failed-replay-sleep", type=float, default=float(CONFIG.get("failed_replay_sleep", 30.0)))
    replay_enabled = bool(CONFIG.get("failed_replay_enabled", False))
    parser.add_argument("--failed-replay", dest="no_failed_replay", action="store_false")
    parser.add_argument("--no-failed-replay", dest="no_failed_replay", action="store_true")
    parser.set_defaults(no_failed_replay=not replay_enabled)


def cmd_verify_one(args: argparse.Namespace) -> int:
    args.stage_image = str(Path(args.qr_image).resolve())
    args.stop_at = "result"
    args.return_to_entry = True
    try:
        with PhoneLock(task_id=getattr(args, "invoice_task_id", None), wait_seconds=args.phone_lock_wait_seconds):
            try:
                code = run_agent(args)
            finally:
                if args.cleanup_phone_album and not args.keep_phone_album and not is_cancelled(getattr(args, "cancel_file", None)):
                    cleanup_phone_album()
        out_dir = Path(args.output_dir).resolve()
        summary = load_single_summary(out_dir / "summary.json")
        result = build_single_result(args, summary)
        archive_path = write_single_outputs(out_dir, result)
        process_archive = archive_process_pages(out_dir)
        if process_archive:
            summary["process_trace_archive"] = str(process_archive)
            (out_dir / "summary.json").write_text(
                json.dumps(summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        emit_agent(
            "result",
            "最终交付物已生成：Excel 清单、核验结果截图目录和结果压缩包。",
            Excel=out_dir / "发票核验结果清单.xlsx",
            截图目录=out_dir / "validation_screenshots",
            压缩包=archive_path,
        )
        return code
    except CancelledError:
        emit_agent("result", "发票查验已取消。")
        return 130


def cmd_verify_batch(args: argparse.Namespace) -> int:
    args.batch_input_dir = args.input_dir
    args.stop_at = "result"
    args.return_to_entry = True
    try:
        with PhoneLock(task_id=getattr(args, "invoice_task_id", None), wait_seconds=args.phone_lock_wait_seconds):
            return run_batch(args)
    except CancelledError:
        emit_agent("result", "发票查验批处理已取消。")
        return 130


def cmd_submit(args: argparse.Namespace) -> int:
    payload = submit_queue_task(args)
    print(json.dumps(payload, ensure_ascii=False))
    return 0


def cmd_queue_status(_: argparse.Namespace) -> int:
    print(json.dumps(queue_status_payload(), ensure_ascii=False, indent=2))
    return 0


def cmd_cancel(args: argparse.Namespace) -> int:
    payload = request_cancel(args.invoice_task_id)
    print(json.dumps(payload, ensure_ascii=False))
    return 0 if payload.get("success") else 2


def cmd_worker(args: argparse.Namespace) -> int:
    processed = 0
    while True:
        with queue_connect() as conn:
            task = fetch_next_queue_task(conn)
        if task is None:
            if args.once:
                print(json.dumps({"success": True, "status": "idle", "processed": processed}, ensure_ascii=False))
                return 0
            sleep_with_cancel(args.poll_seconds)
            continue

        task_id = str(task["invoice_task_id"])
        cancel_file = cancel_file_for_task(task_id)
        if cancel_file.is_file():
            now = utc_now()
            update_queue_task(task_id, status="cancelled", cancelled_at=now, finished_at=now)
            continue

        emit_progress("invoice queue task claimed", invoice_task_id=task_id, title=task.get("title"))
        worker_args = argparse.Namespace(**vars(args))
        worker_args.input_dir = task["input_dir"]
        worker_args.output_dir = task["output_dir"]
        worker_args.batch_input_dir = task["input_dir"]
        worker_args.invoice_task_id = task_id
        worker_args.cancel_file = str(cancel_file)
        worker_args.stop_at = "result"
        worker_args.return_to_entry = True
        worker_args.keep_phone_album = bool(getattr(args, "keep_phone_album", False))
        try:
            with PhoneLock(task_id=task_id, wait_seconds=args.phone_lock_wait_seconds):
                code = run_batch(worker_args)
            summary_path = str(Path(worker_args.output_dir).resolve() / "batch_summary.json")
            if code == 0:
                status = "completed"
            elif code == 130:
                status = "cancelled"
            else:
                status = "partial_failed"
            update_queue_task(task_id, status=status, finished_at=utc_now(), summary_path=summary_path)
        except CancelledError:
            update_queue_task(
                task_id,
                status="cancelled",
                cancelled_at=utc_now(),
                finished_at=utc_now(),
                error_message="cancelled_by_user",
            )
        except SystemExit as exc:
            update_queue_task(
                task_id,
                status="failed",
                finished_at=utc_now(),
                error_message=f"SystemExit: {exc}",
            )
            if not args.keep_going:
                raise
        except Exception as exc:  # noqa: BLE001 - worker must keep queue observable
            update_queue_task(
                task_id,
                status="failed",
                finished_at=utc_now(),
                error_message=f"{type(exc).__name__}: {exc}",
            )
            if not args.keep_going:
                raise
        processed += 1
        if args.once:
            print(json.dumps({"success": True, "status": "processed", "processed": processed}, ensure_ascii=False))
            return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Invoice check phone-use CLI service for AgentD.")
    sub = parser.add_subparsers(dest="command", required=True)

    health = sub.add_parser("health", help="Print service configuration.")
    health.set_defaults(func=cmd_health)

    one = sub.add_parser("verify-one", help="Verify one QR image through the Android phone.")
    one.add_argument("--qr-image", required=True)
    one.add_argument("--output-dir", required=True)
    add_runtime_args(one)
    one.set_defaults(func=cmd_verify_one)

    batch = sub.add_parser("verify-batch", help="Verify a directory of QR images.")
    batch.add_argument("--input-dir", required=True, help="Directory containing extracted QR images.")
    batch.add_argument("--output-dir", required=True)
    add_runtime_args(batch)
    add_batch_policy_args(batch)
    batch.set_defaults(func=cmd_verify_batch)

    submit = sub.add_parser("submit", help="Submit a batch to the invoice-check phone queue.")
    submit.add_argument("--input-dir", required=True, help="Directory containing extracted QR images.")
    submit.add_argument("--output-dir", required=True)
    submit.add_argument("--invoice-task-id")
    submit.add_argument("--session-id")
    submit.add_argument("--user-id")
    submit.add_argument("--title")
    submit.add_argument("--priority", type=int, default=100)
    submit.set_defaults(func=cmd_submit)

    worker = sub.add_parser("worker", help="Run the invoice-check phone queue worker.")
    add_runtime_args(worker)
    add_batch_policy_args(worker)
    worker.add_argument("--once", action="store_true", help="Process at most one queued task, then exit.")
    worker.add_argument("--poll-seconds", type=float, default=3.0)
    worker.add_argument("--keep-going", action=argparse.BooleanOptionalAction, default=True)
    worker.set_defaults(func=cmd_worker)

    status = sub.add_parser("queue-status", help="Print invoice-check phone queue status.")
    status.set_defaults(func=cmd_queue_status)

    cancel = sub.add_parser("cancel", help="Cancel a queued or running invoice-check task.")
    cancel.add_argument("invoice_task_id")
    cancel.set_defaults(func=cmd_cancel)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
